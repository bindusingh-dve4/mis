import os
import bcrypt
from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from sqlalchemy import event
from email_service import email_service
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import logging
import pytz
import calendar
from config import (
    UPLOAD_WINDOW_START_DAY, UPLOAD_WINDOW_END_DAY,
    UPLOAD_WINDOW_REMINDER_DAY, UPLOAD_LOCK_DAY,
    UPLOAD_WINDOW_OPEN_HOUR, UPLOAD_WINDOW_OPEN_MINUTE,
    UPLOAD_WINDOW_REMINDER_HOUR, UPLOAD_WINDOW_REMINDER_MINUTE,
    UPLOAD_WINDOW_LOCK_HOUR, UPLOAD_WINDOW_LOCK_MINUTE,
    SUPERVISOR_APPROVAL_START_DAY, SUPERVISOR_APPROVAL_HOUR, SUPERVISOR_APPROVAL_MINUTE
)

# Set IST timezone
IST = pytz.timezone('Asia/Kolkata')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'dev-secret-key-change-in-production')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///mis_config.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

db = SQLAlchemy(app)

class Role(db.Model):
    __tablename__ = 'roles'
    RoleID = db.Column(db.Integer, primary_key=True)
    RoleName = db.Column(db.String(50), unique=True, nullable=False)
    users = db.relationship('User', backref='role', lazy=True)

class Company(db.Model):
    __tablename__ = 'companies'
    CompanyID = db.Column(db.Integer, primary_key=True)
    CompanyName = db.Column(db.String(100), unique=True, nullable=False)
    ActiveFlag = db.Column(db.Boolean, default=True)

class Department(db.Model):
    __tablename__ = 'departments'
    DeptID = db.Column(db.Integer, primary_key=True)
    DeptName = db.Column(db.String(100), unique=True, nullable=False)
    ActiveFlag = db.Column(db.Boolean, default=True)
    users = db.relationship('User', backref='department', lazy=True)
    uploads = db.relationship('MISUpload', backref='department', lazy=True)
    templates = db.relationship('Template', backref='department', lazy=True)

class FinancialYear(db.Model):
    __tablename__ = 'financial_years'
    FYID = db.Column(db.Integer, primary_key=True)
    FYName = db.Column(db.String(50), unique=True, nullable=False)
    StartDate = db.Column(db.Date, nullable=False)
    EndDate = db.Column(db.Date, nullable=False)
    ActiveFlag = db.Column(db.Boolean, default=False)
    uploads = db.relationship('MISUpload', backref='financial_year', lazy=True)

class User(db.Model):
    __tablename__ = 'users'
    UserID = db.Column(db.Integer, primary_key=True)
    EmpID = db.Column(db.String(20), unique=True, nullable=False, index=True)
    Username = db.Column(db.String(50), nullable=True)
    PasswordHash = db.Column(db.String(255), nullable=False)
    Email = db.Column(db.String(100), unique=True, nullable=False)
    DepartmentID = db.Column(db.Integer, db.ForeignKey('departments.DeptID'), nullable=False)
    RoleID = db.Column(db.Integer, db.ForeignKey('roles.RoleID'), nullable=False)
    IsActive = db.Column(db.Boolean, default=True)
    FailedLoginAttempts = db.Column(db.Integer, default=0)
    LastFailedLogin = db.Column(db.DateTime, nullable=True)
    AccountLockedUntil = db.Column(db.DateTime, nullable=True)
    PasswordLastChanged = db.Column(db.DateTime, default=lambda: datetime.now(IST))
    PasswordExpiryDate = db.Column(db.DateTime, default=lambda: datetime.now(IST) + timedelta(days=90))
    CreatedAt = db.Column(db.DateTime, default=lambda: datetime.now(IST))
    UpdatedAt = db.Column(db.DateTime, default=lambda: datetime.now(IST), onupdate=lambda: datetime.now(IST))
    uploads = db.relationship('MISUpload', foreign_keys='MISUpload.UploadedBy', backref='uploader', lazy=True)

class MISUpload(db.Model):
    __tablename__ = 'mis_uploads'
    UploadID = db.Column(db.Integer, primary_key=True)
    UploadCode = db.Column(db.String(50), unique=True, nullable=True)
    DepartmentID = db.Column(db.Integer, db.ForeignKey('departments.DeptID'), nullable=False)
    MonthID = db.Column(db.Integer, nullable=False)
    FYID = db.Column(db.Integer, db.ForeignKey('financial_years.FYID'), nullable=False)
    UploadedBy = db.Column(db.Integer, db.ForeignKey('users.UserID'), nullable=False)
    UploadDate = db.Column(db.DateTime, default=lambda: datetime.now(IST))
    FilePath = db.Column(db.String(255), nullable=False)
    FileCheck = db.Column(db.String(50), default='Not Validated')
    Status = db.Column(db.String(50), default='In Review')
    IsModified = db.Column(db.Boolean, default=False)
    IsCancelled = db.Column(db.Boolean, default=False)
    SupervisorApproved = db.Column(db.Boolean, default=False)
    SupervisorApprovedBy = db.Column(db.Integer, db.ForeignKey('users.UserID'), nullable=True)
    SupervisorApprovedDate = db.Column(db.DateTime, nullable=True)

class Template(db.Model):
    __tablename__ = 'templates'
    TemplateID = db.Column(db.Integer, primary_key=True)
    DepartmentID = db.Column(db.Integer, db.ForeignKey('departments.DeptID'), nullable=False)
    FilePath = db.Column(db.String(255), nullable=False)
    UploadDate = db.Column(db.DateTime, default=lambda: datetime.now(IST))

class ConsolidatedMIS(db.Model):
    __tablename__ = 'consolidated_mis'
    ConsolidatedMISID = db.Column(db.Integer, primary_key=True)
    SupervisorID = db.Column(db.Integer, db.ForeignKey('users.UserID'), nullable=False)
    FYID = db.Column(db.Integer, db.ForeignKey('financial_years.FYID'), nullable=False)
    MonthID = db.Column(db.Integer, nullable=False)
    UploadedHODMISIDs = db.Column(db.String(500), nullable=True)
    ConsolidatedFilePath = db.Column(db.String(255), nullable=False)
    Status = db.Column(db.String(50), default='Pending Review')
    CreatedDate = db.Column(db.DateTime, default=lambda: datetime.now(IST))
    ApprovedDate = db.Column(db.DateTime, nullable=True)
    ApprovedBy = db.Column(db.Integer, db.ForeignKey('users.UserID'), nullable=True)
    supervisor = db.relationship('User', foreign_keys=[SupervisorID], backref='consolidated_uploads')
    financial_year = db.relationship('FinancialYear', backref='consolidated_mis')

class Notification(db.Model):
    __tablename__ = 'notifications'
    NotificationID = db.Column(db.Integer, primary_key=True)
    UserID = db.Column(db.Integer, db.ForeignKey('users.UserID'), nullable=False)
    Title = db.Column(db.String(200), nullable=False)
    Message = db.Column(db.Text, nullable=False)
    NotificationType = db.Column(db.String(50), nullable=False)
    RelatedUploadID = db.Column(db.Integer, db.ForeignKey('mis_uploads.UploadID'), nullable=True)
    RelatedConsolidatedID = db.Column(db.Integer, db.ForeignKey('consolidated_mis.ConsolidatedMISID'), nullable=True)
    IsRead = db.Column(db.Boolean, default=False)
    CreatedDate = db.Column(db.DateTime, default=lambda: datetime.now(IST))
    user = db.relationship('User', backref='notifications')

def hash_password(password):
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password, hashed):
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def generate_mis_code(department_id):
    """Generate unique MIS code in format: MIS+DPT+[code]"""
    dept = Department.query.get(department_id)
    if not dept:
        return None
    
    dept_code = dept.DeptName[:3].upper()
    
    # Count existing uploads for this department to generate sequential code
    upload_count = MISUpload.query.filter_by(DepartmentID=department_id).count()
    sequential_code = str(upload_count + 1).zfill(6)
    
    return f"MIS{dept_code}{sequential_code}"

def create_notification(user_id, title, message, notif_type, upload_id=None, consolidated_id=None):
    """Create in-app notification for a user"""
    notification = Notification(
        UserID=user_id,
        Title=title,
        Message=message,
        NotificationType=notif_type,
        RelatedUploadID=upload_id,
        RelatedConsolidatedID=consolidated_id
    )
    db.session.add(notification)
    db.session.commit()

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.IsActive:
            session.clear()
            flash('Your session has expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        if user.role.RoleName != 'Admin':
            flash('Access denied. Admin privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def hod_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.IsActive:
            session.clear()
            flash('Your session has expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        if user.role.RoleName not in ['Admin', 'HOD']:
            flash('Access denied. HOD privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def management_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.IsActive:
            session.clear()
            flash('Your session has expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        if user.role.RoleName != 'Management':
            flash('Access denied. Management privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def supervisor_required(f):
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.IsActive:
            session.clear()
            flash('Your session has expired. Please log in again.', 'error')
            return redirect(url_for('login'))
        if user.role.RoleName != 'Supervisor':
            flash('Access denied. Supervisor privileges required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def check_upload_window():
    today = date.today()
    if UPLOAD_WINDOW_START_DAY <= today.day <= UPLOAD_WINDOW_END_DAY:
        days_remaining = UPLOAD_WINDOW_END_DAY - today.day
        if days_remaining == 0:
            return True, f"Upload window is open ({UPLOAD_WINDOW_START_DAY}st-{UPLOAD_WINDOW_END_DAY}th of the month). Last day to upload!"
        elif days_remaining == 1:
            return True, f"Upload window is open ({UPLOAD_WINDOW_START_DAY}st-{UPLOAD_WINDOW_END_DAY}th of the month). 1 day remaining."
        else:
            return True, f"Upload window is open ({UPLOAD_WINDOW_START_DAY}st-{UPLOAD_WINDOW_END_DAY}th of the month). {days_remaining} days remaining."
    elif today.day > UPLOAD_WINDOW_END_DAY:
        return False, f"Upload window is closed. The upload window opens on the {UPLOAD_WINDOW_START_DAY}st of next month."
    else:
        return False, f"Upload window opens on the {UPLOAD_WINDOW_START_DAY}st of the month."

def validate_excel_file(file_path):
    """Validate Excel file structure and content"""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path)
        
        if len(workbook.sheetnames) == 0:
            return False, "Excel file has no sheets."
        
        sheet = workbook.active
        if sheet.max_row < 2:
            return False, "Excel file appears to be empty (no data rows)."
        
        if sheet.max_column < 1:
            return False, "Excel file has no columns."
        
        return True, "File validation successful."
    except Exception as e:
        return False, f"File validation error: {str(e)}"

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        emp_id = request.form.get('emp_id')
        password = request.form.get('password')
        
        user = User.query.filter_by(EmpID=emp_id).first()
        
        if not user:
            flash('Invalid Employee ID or password.', 'error')
            return render_template('login.html')
        
        # Check if account is locked
        if user.AccountLockedUntil:
            locked_until = user.AccountLockedUntil
            # Make timezone-aware if naive
            if locked_until.tzinfo is None:
                locked_until = locked_until.replace(tzinfo=IST)
            if locked_until > datetime.now(IST):
                remaining = (locked_until - datetime.now(IST)).seconds // 60
                flash(f'Account locked due to multiple failed login attempts. Try again in {remaining} minutes.', 'error')
                return render_template('login.html')
            else:
                # Reset lock if time has passed
                user.AccountLockedUntil = None
                user.FailedLoginAttempts = 0
                db.session.commit()
        
        # Check if account is active
        if not user.IsActive:
            flash('Your account is inactive. Please contact administrator.', 'error')
            return render_template('login.html')
        
        # Check password expiry
        if user.PasswordExpiryDate:
            expiry_date = user.PasswordExpiryDate
            # Make timezone-aware if naive
            if expiry_date.tzinfo is None:
                expiry_date = expiry_date.replace(tzinfo=IST)
            if expiry_date < datetime.now(IST):
                flash('Your password has expired. Please contact administrator to reset it.', 'error')
                return render_template('login.html')
        
        # Verify password
        if verify_password(password, user.PasswordHash):
            # Reset failed attempts on successful login
            user.FailedLoginAttempts = 0
            user.LastFailedLogin = None
            db.session.commit()
            
            session['user_id'] = user.UserID
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            # Increment failed attempts
            user.FailedLoginAttempts += 1
            user.LastFailedLogin = datetime.now(IST)
            
            # Lock account after 5 failed attempts
            if user.FailedLoginAttempts >= 5:
                user.AccountLockedUntil = datetime.now(IST) + timedelta(minutes=30)
                db.session.commit()
                flash('Account locked due to 5 failed login attempts. Try again in 30 minutes.', 'error')
            else:
                db.session.commit()
                remaining_attempts = 5 - user.FailedLoginAttempts
                flash(f'Invalid password. {remaining_attempts} attempt(s) remaining before account lockout.', 'error')
    
    return render_template('login.html')

@app.route('/change-password', methods=['POST'])
@login_required
def change_password():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    # Validate current password
    if not verify_password(current_password, user.PasswordHash):
        flash('Current password is incorrect.', 'error')
        return redirect(url_for('dashboard'))
    
    # Validate new password
    if not new_password or len(new_password) < 6:
        flash('New password must be at least 6 characters long.', 'error')
        return redirect(url_for('dashboard'))
    
    # Validate password confirmation
    if new_password != confirm_password:
        flash('New password and confirm password do not match.', 'error')
        return redirect(url_for('dashboard'))
    
    # Update password
    user.PasswordHash = hash_password(new_password)
    user.PasswordLastChanged = datetime.now(IST)
    user.PasswordExpiryDate = datetime.now(IST) + timedelta(days=90)
    db.session.commit()
    
    flash('Password changed successfully! Your new password will expire in 90 days.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))

def analyze_excel_data(file_path):
    """Analyze Excel file and extract key metrics"""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active
        
        # Get total rows and columns
        total_rows = sheet.max_row - 1  # Exclude header
        total_columns = sheet.max_column
        
        # Extract numeric data for analysis
        numeric_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for cell in row:
                if isinstance(cell, (int, float)) and cell is not None:
                    numeric_data.append(cell)
        
        # Calculate metrics
        total_records = total_rows
        total_numeric_values = len(numeric_data)
        avg_value = sum(numeric_data) / len(numeric_data) if numeric_data else 0
        max_value = max(numeric_data) if numeric_data else 0
        min_value = min(numeric_data) if numeric_data else 0
        
        return {
            'total_records': total_records,
            'total_columns': total_columns,
            'total_numeric_values': total_numeric_values,
            'average_value': round(avg_value, 2),
            'max_value': max_value,
            'min_value': min_value,
            'has_data': total_records > 0
        }
    except Exception as e:
        logging.error(f"Error analyzing Excel file {file_path}: {str(e)}")
        return None

@app.route('/dashboard')
@login_required
def dashboard():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
    
    # Get recent uploads based on role
    if user.role.RoleName == 'Admin':
        recent_uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).limit(5).all()
    elif user.role.RoleName == 'Management':
        recent_uploads = ConsolidatedMIS.query.order_by(ConsolidatedMIS.CreatedDate.desc()).limit(5).all()
    elif user.role.RoleName == 'Supervisor':
        recent_uploads = MISUpload.query.filter_by(SupervisorApproved=False, IsCancelled=False).order_by(MISUpload.UploadDate.desc()).limit(5).all()
    elif user.role.RoleName == 'HOD':
        recent_uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).limit(5).all()
    else:
        flash('Invalid role. Please contact administrator.', 'error')
        session.clear()
        return redirect(url_for('login'))
    
    # Check upload window status
    upload_allowed, upload_message = check_upload_window()
    
    # Get HOD count and pending uploads for admin, supervisor and management
    hod_count = 0
    pending_uploads_count = 0
    supervisor_pending_count = 0
    supervisor_pending_uploads = []
    management_pending_consolidated = []
    if user.role.RoleName == 'Admin':
        hod_role = Role.query.filter_by(RoleName='HOD').first()
        if hod_role:
            hod_count = User.query.filter_by(RoleID=hod_role.RoleID, IsActive=True).count()
    if user.role.RoleName == 'Supervisor':
        supervisor_pending_count = MISUpload.query.filter_by(SupervisorApproved=False, Status='In Review', IsCancelled=False).count()
        supervisor_pending_uploads = MISUpload.query.filter_by(SupervisorApproved=False, Status='In Review', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).limit(2).all()
    if user.role.RoleName == 'Management':
        pending_uploads_count = ConsolidatedMIS.query.filter_by(Status='Pending Review').count()
        management_pending_consolidated = ConsolidatedMIS.query.filter_by(Status='Pending Review').order_by(ConsolidatedMIS.CreatedDate.desc()).limit(2).all()
    
    stats = {
        'total_users': User.query.count(),
        'total_depts': Department.query.count(),
        'active_fy': active_fy.FYName if active_fy else 'None',
        'total_uploads': MISUpload.query.count(),
        'recent_uploads': recent_uploads,
        'upload_allowed': upload_allowed,
        'upload_message': upload_message,
        'hod_count': hod_count,
        'pending_uploads_count': pending_uploads_count,
        'supervisor_pending_count': supervisor_pending_count,
        'supervisor_pending_uploads': supervisor_pending_uploads,
        'management_pending_consolidated': management_pending_consolidated,
        'email_configured': email_service.is_configured()
    }
    
    # Calculate last day of current month
    today = date.today()
    last_day_of_month = calendar.monthrange(today.year, today.month)[1]
    
    return render_template('dashboard.html', current_user=user, stats=stats, 
                          upload_window_start=UPLOAD_WINDOW_START_DAY, 
                          upload_window_end=UPLOAD_WINDOW_END_DAY,
                          supervisor_approval_start_day=SUPERVISOR_APPROVAL_START_DAY,
                          supervisor_approval_end_day=last_day_of_month)


@app.route('/my-uploads')
@login_required
def my_uploads():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # HOD can only see their department's uploads
    if user.role.RoleName == 'HOD':
        uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).all()
    else:
        # If not HOD, redirect to reports page
        return redirect(url_for('reports'))
    
    return render_template('my_uploads.html', 
                         current_user=user,
                         uploads=uploads)

@app.route('/reports')
@login_required
def reports():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Filter uploads based on role
    if user.role.RoleName == 'Admin':
        query = MISUpload.query
    elif user.role.RoleName == 'Management':
        query = MISUpload.query
    elif user.role.RoleName == 'Supervisor':
        # Supervisor sees all department records
        query = MISUpload.query
    else:
        # HOD sees only their department records
        query = MISUpload.query.filter_by(DepartmentID=user.DepartmentID)
    
    # Exclude cancelled uploads from all reports
    query = query.filter_by(IsCancelled=False)
    
    # Apply filter parameters
    department_id = request.args.get('department', '')
    fy_id = request.args.get('fy', '')
    status = request.args.get('status', '')
    search_code = request.args.get('search_code', '').strip()
    
    # Search by MIS code if provided
    if search_code:
        query = query.filter_by(UploadCode=search_code)
    else:
        # Apply other filters only if search_code is not provided
        if department_id:
            query = query.filter_by(DepartmentID=int(department_id))
        
        if fy_id:
            query = query.filter_by(FYID=int(fy_id))
        
        if status:
            query = query.filter_by(Status=status)
    
    uploads = query.order_by(MISUpload.UploadDate.desc()).all()
    
    departments = Department.query.filter_by(ActiveFlag=True).all()
    financial_years = FinancialYear.query.all()
    
    return render_template('reports.html', 
                         current_user=user,
                         uploads=uploads,
                         departments=departments,
                         financial_years=financial_years,
                         selected_department=department_id,
                         selected_fy=fy_id,
                         selected_status=status,
                         search_code=search_code,
                         user_role=user.role.RoleName)

@app.route('/approval-queue')
@supervisor_required
def approval_queue():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Supervisor role - get all uploads pending supervisor approval
    # Excludes cancelled uploads (when HOD deletes, upload is removed from approval queue)
    pending_uploads = MISUpload.query.filter_by(SupervisorApproved=False, Status='In Review', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    
    return render_template('approval_queue.html', 
                         current_user=user,
                         pending_uploads=pending_uploads)

@app.route('/management-history')
@management_required
def management_history():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get all MIS uploads with their complete history
    all_uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).all()
    
    # Get filter parameters from query string
    department_id = request.args.get('department_id', '')
    fy_id = request.args.get('fy_id', '')
    status = request.args.get('status', '')
    
    # Apply filters if provided
    query = MISUpload.query
    
    if department_id:
        query = query.filter_by(DepartmentID=int(department_id))
    
    if fy_id:
        query = query.filter_by(FYID=int(fy_id))
    
    if status:
        query = query.filter_by(Status=status)
    
    all_uploads = query.order_by(MISUpload.UploadDate.desc()).all()
    
    departments = Department.query.filter_by(ActiveFlag=True).all()
    financial_years = FinancialYear.query.all()
    
    return render_template('management_history.html', 
                         current_user=user,
                         uploads=all_uploads,
                         departments=departments,
                         financial_years=financial_years,
                         selected_department=department_id,
                         selected_fy=fy_id,
                         selected_status=status)

@app.route('/approved-mis')
@admin_required
def approved_mis():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Admin role - get all approved MIS uploads from all departments
    approved_uploads = MISUpload.query.filter_by(Status='Approved').order_by(MISUpload.UploadDate.desc()).all()
    
    return render_template('approved_mis.html', 
                         current_user=user,
                         uploads=approved_uploads)

@app.route('/view-upload/<int:upload_id>')
@login_required
def view_upload(upload_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions: Admin, Management, and HOD can view uploads
    if user.role.RoleName == 'Management':
        pass  # Management can view all uploads
    elif user.role.RoleName == 'Admin':
        pass  # Admin can view all uploads
    elif user.role.RoleName == 'HOD' and upload.DepartmentID != user.DepartmentID:
        flash('Access denied. You can only view uploads from your department.', 'error')
        return redirect(url_for('reports'))
    else:
        flash('Access denied.', 'error')
        return redirect(url_for('reports'))
    
    return render_template('view_upload.html', 
                         current_user=user,
                         upload=upload)

@app.route('/approve-upload/<int:upload_id>', methods=['POST'])
@management_required
def approve_upload(upload_id):
    user = User.query.get(session['user_id'])
    
    upload = MISUpload.query.get_or_404(upload_id)
    
    upload.Status = 'Approved'
    db.session.commit()
    
    # Send email notification to uploader
    if email_service.is_configured():
        uploader = upload.uploader
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[upload.MonthID]
        
        subject = f"MIS Upload Approved - {upload.department.DeptName} - {month_name} {upload.financial_year.FYName}"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #10b981; color: white; padding: 20px; border-radius: 5px 5px 0 0; }}
                .content {{ background-color: #f9fafb; padding: 30px; border: 1px solid #e5e7eb; }}
                .success-box {{ background-color: #d1fae5; border-left: 4px solid #10b981; padding: 15px; margin: 20px 0; }}
                .footer {{ background-color: #f3f4f6; padding: 15px; text-align: center; border-radius: 0 0 5px 5px; font-size: 12px; color: #6b7280; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2 style="margin: 0;">✓ MIS Upload Approved</h2>
                </div>
                <div class="content">
                    <p>Dear {uploader.Username},</p>
                    
                    <div class="success-box">
                        <strong>Great news!</strong> Your MIS upload has been approved by the administrator.
                    </div>
                    
                    <h3>Upload Details:</h3>
                    <ul>
                        <li><strong>Department:</strong> {upload.department.DeptName}</li>
                        <li><strong>Month:</strong> {month_name}</li>
                        <li><strong>Financial Year:</strong> {upload.financial_year.FYName}</li>
                        <li><strong>Upload Date:</strong> {upload.UploadDate.strftime('%d %b %Y %H:%M')}</li>
                        <li><strong>Approved:</strong> {datetime.now(IST).strftime('%d %b %Y %H:%M:%S IST')}</li>
                        <li><strong>Status:</strong> Approved</li>
                    </ul>
                    
                    <p>Thank you for your timely submission!</p>
                    
                    <p>Best regards,<br>
                    <strong>MIS System Team</strong></p>
                </div>
                <div class="footer">
                    This is an automated notification from the MIS Upload System.
                </div>
            </div>
        </body>
        </html>
        """
        
        text_content = f"""
MIS Upload Approved

Dear {uploader.Username},

✓ Great news! Your MIS upload has been approved by the administrator.

Upload Details:
- Department: {upload.department.DeptName}
- Month: {month_name}
- Financial Year: {upload.financial_year.FYName}
- Upload Date: {upload.UploadDate.strftime('%d %b %Y %H:%M')}
- Approved: {datetime.now(IST).strftime('%d %b %Y %H:%M:%S IST')}
- Status: Approved

Thank you for your timely submission!

Best regards,
MIS System Team

---
This is an automated notification from the MIS Upload System.
        """
        
        email_service.send_email(uploader.Email, subject, html_content, text_content)
    
    flash(f'Upload approved successfully! Notification sent to {upload.uploader.Username}.', 'success')
    return redirect(url_for('approval_queue'))

@app.route('/reject-upload/<int:upload_id>', methods=['POST'])
@management_required
def reject_upload(upload_id):
    user = User.query.get(session['user_id'])
    
    upload = MISUpload.query.get_or_404(upload_id)
    
    upload.Status = 'Rejected'
    db.session.commit()
    
    # Send email notification to uploader
    if email_service.is_configured():
        uploader = upload.uploader
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[upload.MonthID]
        
        subject = f"MIS Upload Rejected - {upload.department.DeptName} - {month_name} {upload.financial_year.FYName}"
        
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background-color: #ef4444; color: white; padding: 20px; border-radius: 5px 5px 0 0; }}
                .content {{ background-color: #f9fafb; padding: 30px; border: 1px solid #e5e7eb; }}
                .warning-box {{ background-color: #fee2e2; border-left: 4px solid #ef4444; padding: 15px; margin: 20px 0; }}
                .footer {{ background-color: #f3f4f6; padding: 15px; text-align: center; border-radius: 0 0 5px 5px; font-size: 12px; color: #6b7280; }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2 style="margin: 0;">✗ MIS Upload Rejected</h2>
                </div>
                <div class="content">
                    <p>Dear {uploader.Username},</p>
                    
                    <div class="warning-box">
                        <strong>Action Required:</strong> Your MIS upload has been rejected by the administrator.
                    </div>
                    
                    <h3>Upload Details:</h3>
                    <ul>
                        <li><strong>Department:</strong> {upload.department.DeptName}</li>
                        <li><strong>Month:</strong> {month_name}</li>
                        <li><strong>Financial Year:</strong> {upload.financial_year.FYName}</li>
                        <li><strong>Upload Date:</strong> {upload.UploadDate.strftime('%d %b %Y %H:%M')}</li>
                        <li><strong>Rejected:</strong> {datetime.now(IST).strftime('%d %b %Y %H:%M:%S IST')}</li>
                        <li><strong>Status:</strong> Rejected</li>
                    </ul>
                    
                    <p>Please review your submission and re-upload the corrected data during the upload window ({UPLOAD_WINDOW_START_DAY}st-{UPLOAD_WINDOW_END_DAY}th of each month).</p>
                    
                    <p>If you have any questions, please contact the administrator.</p>
                    
                    <p>Best regards,<br>
                    <strong>MIS System Team</strong></p>
                </div>
                <div class="footer">
                    This is an automated notification from the MIS Upload System.
                </div>
            </div>
        </body>
        </html>
        """
        
        text_content = f"""
MIS Upload Rejected

Dear {uploader.Username},

✗ Action Required: Your MIS upload has been rejected by the administrator.

Upload Details:
- Department: {upload.department.DeptName}
- Month: {month_name}
- Financial Year: {upload.financial_year.FYName}
- Upload Date: {upload.UploadDate.strftime('%d %b %Y %H:%M')}
- Rejected: {datetime.now(IST).strftime('%d %b %Y %H:%M:%S IST')}
- Status: Rejected

Please review your submission and re-upload the corrected data during the upload window ({UPLOAD_WINDOW_START_DAY}st-{UPLOAD_WINDOW_END_DAY}th of each month).

If you have any questions, please contact the administrator.

Best regards,
MIS System Team

---
This is an automated notification from the MIS Upload System.
        """
        
        email_service.send_email(uploader.Email, subject, html_content, text_content)
    
    flash(f'Upload rejected. Notification sent to {upload.uploader.Username}.', 'warning')
    return redirect(url_for('approval_queue'))

@app.route('/supervisor-uploads')
@supervisor_required
def supervisor_uploads():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Supervisor can now view all pending uploads regardless of date
    pending_uploads = MISUpload.query.filter_by(SupervisorApproved=False, Status='In Review', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    
    # Get supervisor approved and rejected uploads for history table
    # Match the same filtering logic as supervisor_history route for consistency
    approved_by_supervisor = MISUpload.query.filter_by(SupervisorApproved=True).filter(MISUpload.Status.in_(['In Review', 'Approved'])).order_by(MISUpload.SupervisorApprovedDate.desc()).all()
    rejected_by_supervisor = MISUpload.query.filter_by(SupervisorApproved=False, Status='Rejected').order_by(MISUpload.UploadDate.desc()).all()
    
    return render_template('supervisor_uploads.html', current_user=user, pending_uploads=pending_uploads, approved_by_supervisor=approved_by_supervisor, rejected_by_supervisor=rejected_by_supervisor)

@app.route('/supervisor-history')
@supervisor_required
def supervisor_history():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get approved and rejected uploads by supervisor
    approved_uploads = MISUpload.query.filter_by(SupervisorApproved=True).filter(MISUpload.Status.in_(['In Review', 'Approved'])).order_by(MISUpload.SupervisorApprovedDate.desc()).all()
    rejected_uploads = MISUpload.query.filter_by(SupervisorApproved=False, Status='Rejected').order_by(MISUpload.UploadDate.desc()).all()
    
    # Get consolidated MIS uploads from all supervisors (all departments)
    approved_consolidated = ConsolidatedMIS.query.filter_by(Status='Approved').order_by(ConsolidatedMIS.ApprovedDate.desc()).all()
    rejected_consolidated = ConsolidatedMIS.query.filter_by(Status='Rejected').order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    pending_consolidated = ConsolidatedMIS.query.filter_by(Status='Pending Review').order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    
    return render_template('supervisor_history.html', current_user=user, approved_uploads=approved_uploads, rejected_uploads=rejected_uploads, 
                         approved_consolidated=approved_consolidated, rejected_consolidated=rejected_consolidated, pending_consolidated=pending_consolidated)

@app.route('/supervisor-mis-tracking')
@supervisor_required
def supervisor_mis_tracking():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get filter parameters
    current_month = date.today().month
    active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
    
    selected_month = request.args.get('month_id', str(current_month))
    selected_fy = request.args.get('fy_id', str(active_fy.FYID) if active_fy else '1')
    
    # Get all active departments
    all_departments = Department.query.filter_by(ActiveFlag=True).order_by(Department.DeptName).all()
    
    # Build department status list
    department_statuses = []
    submitted_count = 0
    pending_review_count = 0
    not_submitted_count = 0
    
    for dept in all_departments:
        # Get HOD for this department
        hod_role = Role.query.filter_by(RoleName='HOD').first()
        hod = User.query.filter_by(DepartmentID=dept.DeptID, RoleID=hod_role.RoleID, IsActive=True).first() if hod_role else None
        
        # Get MIS upload for this department, month, and FY
        upload = MISUpload.query.filter_by(
            DepartmentID=dept.DeptID,
            MonthID=int(selected_month),
            FYID=int(selected_fy),
            IsCancelled=False
        ).order_by(MISUpload.UploadDate.desc()).first()
        
        # Count statuses
        if upload:
            submitted_count += 1
            if not upload.SupervisorApproved and upload.Status == 'In Review':
                pending_review_count += 1
        else:
            not_submitted_count += 1
        
        department_statuses.append({
            'department_name': dept.DeptName,
            'department_id': dept.DeptID,
            'hod_name': hod.Username if hod else None,
            'hod_emp_id': hod.EmpID if hod else None,
            'upload': upload
        })
    
    financial_years = FinancialYear.query.all()
    selected_fy_obj = FinancialYear.query.get(int(selected_fy)) if selected_fy else active_fy
    
    return render_template('supervisor_mis_tracking.html',
                         current_user=user,
                         department_statuses=department_statuses,
                         total_departments=len(all_departments),
                         submitted_count=submitted_count,
                         pending_review_count=pending_review_count,
                         not_submitted_count=not_submitted_count,
                         financial_years=financial_years,
                         selected_month=selected_month,
                         selected_fy=selected_fy,
                         selected_fy_name=selected_fy_obj.FYName if selected_fy_obj else '')

@app.route('/admin-mis-tracking')
@admin_required
def admin_mis_tracking():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get filter parameters
    current_month = date.today().month
    active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
    
    selected_month = request.args.get('month_id', str(current_month))
    selected_fy = request.args.get('fy_id', str(active_fy.FYID) if active_fy else '1')
    
    # Get all active departments
    all_departments = Department.query.filter_by(ActiveFlag=True).order_by(Department.DeptName).all()
    
    # Build department status list
    department_statuses = []
    submitted_count = 0
    pending_review_count = 0
    not_submitted_count = 0
    approved_count = 0
    rejected_count = 0
    
    for dept in all_departments:
        # Get HOD for this department
        hod_role = Role.query.filter_by(RoleName='HOD').first()
        hod = User.query.filter_by(DepartmentID=dept.DeptID, RoleID=hod_role.RoleID, IsActive=True).first() if hod_role else None
        
        # Get MIS upload for this department, month, and FY
        upload = MISUpload.query.filter_by(
            DepartmentID=dept.DeptID,
            MonthID=int(selected_month),
            FYID=int(selected_fy),
            IsCancelled=False
        ).order_by(MISUpload.UploadDate.desc()).first()
        
        # Count statuses
        if upload:
            submitted_count += 1
            if upload.Status == 'In Review':
                pending_review_count += 1
            elif upload.Status == 'Approved':
                approved_count += 1
            elif upload.Status == 'Rejected':
                rejected_count += 1
        else:
            not_submitted_count += 1
        
        department_statuses.append({
            'department_name': dept.DeptName,
            'department_id': dept.DeptID,
            'hod_name': hod.Username if hod else None,
            'hod_emp_id': hod.EmpID if hod else None,
            'upload': upload
        })
    
    financial_years = FinancialYear.query.all()
    selected_fy_obj = FinancialYear.query.get(int(selected_fy)) if selected_fy else active_fy
    
    return render_template('admin_mis_tracking.html',
                         current_user=user,
                         department_statuses=department_statuses,
                         total_departments=len(all_departments),
                         submitted_count=submitted_count,
                         pending_review_count=pending_review_count,
                         not_submitted_count=not_submitted_count,
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         financial_years=financial_years,
                         selected_month=selected_month,
                         selected_fy=selected_fy,
                         selected_fy_name=selected_fy_obj.FYName if selected_fy_obj else '')

@app.route('/management-mis-tracking')
@management_required
def management_mis_tracking():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get filter parameters
    current_month = date.today().month
    active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
    
    selected_month = request.args.get('month_id', str(current_month))
    selected_fy = request.args.get('fy_id', str(active_fy.FYID) if active_fy else '1')
    
    # Get all active departments
    all_departments = Department.query.filter_by(ActiveFlag=True).order_by(Department.DeptName).all()
    
    # Build department status list
    department_statuses = []
    submitted_count = 0
    pending_review_count = 0
    not_submitted_count = 0
    approved_count = 0
    rejected_count = 0
    
    for dept in all_departments:
        # Get HOD for this department
        hod_role = Role.query.filter_by(RoleName='HOD').first()
        hod = User.query.filter_by(DepartmentID=dept.DeptID, RoleID=hod_role.RoleID, IsActive=True).first() if hod_role else None
        
        # Get MIS upload for this department, month, and FY
        upload = MISUpload.query.filter_by(
            DepartmentID=dept.DeptID,
            MonthID=int(selected_month),
            FYID=int(selected_fy),
            IsCancelled=False
        ).order_by(MISUpload.UploadDate.desc()).first()
        
        # Count statuses
        if upload:
            submitted_count += 1
            if upload.Status == 'In Review':
                pending_review_count += 1
            elif upload.Status == 'Approved':
                approved_count += 1
            elif upload.Status == 'Rejected':
                rejected_count += 1
        else:
            not_submitted_count += 1
        
        department_statuses.append({
            'department_name': dept.DeptName,
            'department_id': dept.DeptID,
            'hod_name': hod.Username if hod else None,
            'hod_emp_id': hod.EmpID if hod else None,
            'upload': upload
        })
    
    financial_years = FinancialYear.query.all()
    selected_fy_obj = FinancialYear.query.get(int(selected_fy)) if selected_fy else active_fy
    
    return render_template('management_mis_tracking.html',
                         current_user=user,
                         department_statuses=department_statuses,
                         total_departments=len(all_departments),
                         submitted_count=submitted_count,
                         pending_review_count=pending_review_count,
                         not_submitted_count=not_submitted_count,
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         financial_years=financial_years,
                         selected_month=selected_month,
                         selected_fy=selected_fy,
                         selected_fy_name=selected_fy_obj.FYName if selected_fy_obj else '')

@app.route('/consolidated-mis-dashboard')
@management_required
def consolidated_mis_dashboard():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get filter parameters
    fy_id = request.args.get('fy_id', '')
    month_id = request.args.get('month_id', '')
    status = request.args.get('status', '')
    
    # Base query - get ALL consolidated reports
    query = ConsolidatedMIS.query
    
    # Apply filters only if specified
    if fy_id:
        query = query.filter_by(FYID=int(fy_id))
    
    if month_id:
        query = query.filter_by(MonthID=int(month_id))
    
    if status:
        query = query.filter_by(Status=status)
    
    # Get all consolidated reports sorted by most recent first
    consolidated_reports = query.order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    
    # Calculate statistics based on consolidated reports
    all_consolidated = ConsolidatedMIS.query.all()
    stats = {
        'total_consolidated': len(all_consolidated),
        'pending_count': ConsolidatedMIS.query.filter_by(Status='Pending Review').count(),
        'approved_count': ConsolidatedMIS.query.filter_by(Status='Approved').count(),
        'rejected_count': ConsolidatedMIS.query.filter_by(Status='Rejected').count()
    }
    
    # Prepare chart data based on consolidated reports
    chart_data = []
    for report in consolidated_reports:
        chart_data.append({
            'MonthID': report.MonthID,
            'Status': report.Status
        })
    
    financial_years = FinancialYear.query.all()
    
    return render_template('consolidated_mis_dashboard.html', 
                         current_user=user,
                         consolidated_reports=consolidated_reports,
                         financial_years=financial_years,
                         stats=stats,
                         chart_data=chart_data,
                         selected_fy=fy_id,
                         selected_month=month_id,
                         selected_status=status)

@app.route('/management-consolidated-reports')
@management_required
def management_consolidated_reports():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get filter parameters
    fy_id = request.args.get('fy_id', '')
    month_id = request.args.get('month_id', '')
    status = request.args.get('status', '')
    
    # Get individual MIS reports (HOD uploads) with filters
    individual_query = MISUpload.query.filter_by(IsCancelled=False)
    
    if fy_id:
        individual_query = individual_query.filter_by(FYID=int(fy_id))
    
    if month_id:
        individual_query = individual_query.filter_by(MonthID=int(month_id))
    
    if status:
        individual_query = individual_query.filter_by(Status=status)
    
    individual_reports = individual_query.order_by(MISUpload.UploadDate.desc()).all()
    
    # Calculate statistics based on individual reports
    all_individual = MISUpload.query.filter_by(IsCancelled=False).all()
    stats = {
        'total_reports': len(all_individual),
        'pending_count': MISUpload.query.filter_by(Status='In Review', IsCancelled=False).count(),
        'approved_count': MISUpload.query.filter_by(Status='Approved', IsCancelled=False).count(),
        'rejected_count': MISUpload.query.filter_by(Status='Rejected', IsCancelled=False).count()
    }
    
    # Prepare chart data based on individual reports
    chart_data = []
    for report in individual_reports:
        chart_data.append({
            'MonthID': report.MonthID,
            'Status': report.Status
        })
    
    financial_years = FinancialYear.query.all()
    
    return render_template('management_consolidated_reports.html', 
                         current_user=user,
                         individual_reports=individual_reports,
                         financial_years=financial_years,
                         stats=stats,
                         chart_data=chart_data,
                         selected_fy=fy_id,
                         selected_month=month_id,
                         selected_status=status)

@app.route('/management-individual-detail-month/<int:month_id>')
@management_required
def management_individual_detail_month(month_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get individual reports for the selected month
    reports = MISUpload.query.filter_by(MonthID=month_id, IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[month_id] if 1 <= month_id <= 12 else f'Month {month_id}'
    
    return render_template('individual_month_detail.html', current_user=user, reports=reports, month_name=month_name, month_id=month_id)

@app.route('/management-individual-detail-fy/<int:fy_id>')
@management_required
def management_individual_detail_fy(fy_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    fy = FinancialYear.query.get_or_404(fy_id)
    reports = MISUpload.query.filter_by(FYID=fy_id, IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    
    return render_template('individual_fy_detail.html', current_user=user, reports=reports, fy=fy)

@app.route('/consolidated-detail-month/<int:month_id>')
@management_required
def consolidated_detail_month(month_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get consolidated reports for the selected month
    reports = ConsolidatedMIS.query.filter_by(MonthID=month_id).order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[month_id] if 1 <= month_id <= 12 else f'Month {month_id}'
    
    return render_template('consolidated_month_detail.html', current_user=user, reports=reports, month_name=month_name, month_id=month_id)

@app.route('/consolidated-detail-fy/<int:fy_id>')
@management_required
def consolidated_detail_fy(fy_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    fy = FinancialYear.query.get_or_404(fy_id)
    reports = ConsolidatedMIS.query.filter_by(FYID=fy_id).order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    
    return render_template('consolidated_fy_detail.html', current_user=user, reports=reports, fy=fy)

@app.route('/management-consolidated-history')
@management_required
def management_consolidated_history():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Get approved and rejected consolidated MIS by management
    approved_consolidated = ConsolidatedMIS.query.filter_by(Status='Approved').order_by(ConsolidatedMIS.ApprovedDate.desc()).all()
    rejected_consolidated = ConsolidatedMIS.query.filter_by(Status='Rejected').order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    
    return render_template('management_consolidated_history.html', current_user=user, approved_consolidated=approved_consolidated, rejected_consolidated=rejected_consolidated)

@app.route('/view-supervisor-consolidated-mis/<int:consolidated_id>')
@supervisor_required
def view_supervisor_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    
    # Ensure supervisor can only view their own consolidated MIS
    if consolidated.SupervisorID != user.UserID:
        flash('Access denied.', 'error')
        return redirect(url_for('supervisor_history'))
    
    hod_uploads = MISUpload.query.filter(MISUpload.UploadID.in_([int(x) for x in consolidated.UploadedHODMISIDs.split(',')])).all() if consolidated.UploadedHODMISIDs else []
    
    # Get approver info
    approver = User.query.get(consolidated.ApprovedBy) if consolidated.ApprovedBy else None
    
    return render_template('view_supervisor_consolidated_mis.html', current_user=user, consolidated=consolidated, hod_uploads=hod_uploads, approver=approver)

@app.route('/view-hod-upload/<int:upload_id>')
@supervisor_required
def view_hod_upload(upload_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    upload = MISUpload.query.get_or_404(upload_id)
    return render_template('view_hod_upload.html', current_user=user, upload=upload)

@app.route('/approve-hod-upload/<int:upload_id>', methods=['POST'])
@supervisor_required
def approve_hod_upload(upload_id):
    user = User.query.get(session['user_id'])
    
    # Supervisor can now approve uploads regardless of date
    upload = MISUpload.query.get_or_404(upload_id)
    upload.SupervisorApproved = True
    upload.SupervisorApprovedBy = user.UserID
    upload.SupervisorApprovedDate = datetime.now(IST)
    db.session.commit()
    
    # Create in-app notification for HOD
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[upload.MonthID]
    create_notification(upload.UploadedBy, 'MIS Approved by Supervisor', f'Your MIS upload {upload.UploadCode} for {month_name} has been approved by the Supervisor and is pending Management review.', 'approval', upload_id=upload_id)
    
    if email_service.is_configured():
        uploader = upload.uploader
        subject = f"MIS Upload Approved by Supervisor - {upload.department.DeptName} - {month_name}"
        html_content = f"""<!DOCTYPE html><html><head><style>body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }} .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }} .header {{ background-color: #3b82f6; color: white; padding: 20px; border-radius: 5px 5px 0 0; }} .content {{ background-color: #f9fafb; padding: 30px; border: 1px solid #e5e7eb; }} .success-box {{ background-color: #dbeafe; border-left: 4px solid #3b82f6; padding: 15px; margin: 20px 0; }}</style></head><body><div class="container"><div class="header"><h2>MIS Upload Approved - Supervisor Review</h2></div><div class="content"><p>Dear {uploader.Username},</p><div class="success-box"><strong>Your MIS upload has been approved by the Supervisor!</strong><br>It is now pending Management's final review.</div><p><strong>Details:</strong> {upload.department.DeptName} | {month_name} {upload.financial_year.FYName} | Code: {upload.UploadCode}</p><p>You will be notified once Management completes their review.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></body></html>"""
        email_service.send_email(uploader.Email, subject, html_content)
    
    flash('HOD MIS approved successfully!', 'success')
    return redirect(url_for('supervisor_uploads'))

@app.route('/reject-hod-upload/<int:upload_id>', methods=['POST'])
@supervisor_required
def reject_hod_upload(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    upload.Status = 'Rejected'
    db.session.commit()
    
    # Create in-app notification for HOD
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[upload.MonthID]
    create_notification(upload.UploadedBy, 'MIS Rejected by Supervisor', f'Your MIS upload {upload.UploadCode} for {month_name} has been rejected by the Supervisor. Please review and resubmit.', 'rejection', upload_id=upload_id)
    
    if email_service.is_configured():
        uploader = upload.uploader
        subject = f"MIS Upload Rejected - {upload.department.DeptName} - {month_name}"
        html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #ef4444; color: white; padding: 20px; border-radius: 5px 5px 0 0;"><h2>MIS Upload Rejected</h2></div><div style="background-color: white; padding: 30px;"><p>Dear {uploader.Username},</p><p>Your MIS upload for {upload.department.DeptName} ({month_name}) has been rejected by the Supervisor.</p><p>Please review and resubmit your MIS upload.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
        email_service.send_email(uploader.Email, subject, html_content)
    
    flash('HOD MIS rejected. Notification sent to uploader.', 'warning')
    return redirect(url_for('supervisor_uploads'))

@app.route('/prepare-consolidated-mis')
@supervisor_required
def prepare_consolidated_mis():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Supervisor can now prepare consolidated MIS anytime (no date restriction)
    # Show all supervisor-approved uploads (including those uploaded by Admin)
    all_approved_uploads = MISUpload.query.filter_by(SupervisorApproved=True, IsCancelled=False).filter(MISUpload.Status.in_(['In Review', 'Approved'])).order_by(MISUpload.UploadDate.desc()).all()
    
    # Get all consolidated MIS records to check which uploads are already included
    all_consolidated = ConsolidatedMIS.query.all()
    already_included_upload_ids = set()
    
    for consolidated in all_consolidated:
        if consolidated.UploadedHODMISIDs:
            upload_ids = [int(x) for x in consolidated.UploadedHODMISIDs.split(',')]
            already_included_upload_ids.update(upload_ids)
    
    # Filter out uploads that are already included in any consolidated MIS
    approved_uploads = [upload for upload in all_approved_uploads if upload.UploadID not in already_included_upload_ids]
    
    # Get all uploads with management approval/rejection status
    in_review_uploads = MISUpload.query.filter_by(Status='In Review', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    approved_by_management = MISUpload.query.filter_by(Status='Approved', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    rejected_by_management = MISUpload.query.filter_by(Status='Rejected', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    
    return render_template('prepare_consolidated_mis.html', current_user=user, approved_uploads=approved_uploads, in_review_uploads=in_review_uploads, approved_by_management=approved_by_management, rejected_by_management=rejected_by_management)

@app.route('/upload-consolidated-mis', methods=['POST'])
@supervisor_required
def upload_consolidated_mis():
    user = User.query.get(session['user_id'])
    
    # Supervisor can now upload consolidated MIS anytime (no date restriction)
    if 'consolidated_file' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('prepare_consolidated_mis'))
    
    file = request.files['consolidated_file']
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('prepare_consolidated_mis'))
    
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        flash('Only .xls or .xlsx files are allowed.', 'error')
        return redirect(url_for('prepare_consolidated_mis'))
    
    selected_uploads = request.form.getlist('selected_uploads')
    if not selected_uploads:
        flash('Please select at least one HOD MIS upload.', 'error')
        return redirect(url_for('prepare_consolidated_mis'))
    
    active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
    current_month = date.today().month
    
    # Check for duplicate consolidated MIS for the same period
    existing_consolidated = ConsolidatedMIS.query.filter_by(
        FYID=active_fy.FYID if active_fy else 1,
        MonthID=current_month
    ).first()
    
    if existing_consolidated:
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        flash(f'Consolidated MIS already exists for {month_names[current_month]} {active_fy.FYName if active_fy else ""}. Cannot create duplicate consolidated MIS for the same period.', 'error')
        return redirect(url_for('prepare_consolidated_mis'))
    
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'ConsolidatedMIS', active_fy.FYName if active_fy else 'General')
    os.makedirs(upload_dir, exist_ok=True)
    
    filename = secure_filename(f"ConsolidatedMIS_{current_month:02d}_{active_fy.FYName if active_fy else 'General'}_{file.filename}")
    filepath = os.path.join(upload_dir, filename)
    file.save(filepath)
    
    is_valid, validation_message = validate_excel_file(filepath)
    if not is_valid:
        os.remove(filepath)
        flash(f'Validation Error: {validation_message}', 'error')
        return redirect(url_for('prepare_consolidated_mis'))
    
    consolidated = ConsolidatedMIS(SupervisorID=user.UserID, FYID=active_fy.FYID if active_fy else 1, MonthID=current_month, UploadedHODMISIDs=','.join(selected_uploads), ConsolidatedFilePath=filepath, Status='Pending Review')
    db.session.add(consolidated)
    db.session.commit()
    
    # Send email notifications to all Management users
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[current_month]
    if email_service.is_configured():
        management_role = Role.query.filter_by(RoleName='Management').first()
        if management_role:
            management_users = User.query.filter_by(RoleID=management_role.RoleID, IsActive=True).all()
            for mgmt_user in management_users:
                subject = f"New Consolidated MIS for Review - {month_name}"
                html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #f97316; color: white; padding: 20px; border-radius: 5px 5px 0 0;"><h2>New Consolidated MIS Submitted</h2></div><div style="background-color: white; padding: 30px;"><p>Dear {mgmt_user.Username},</p><p>A new consolidated MIS for {month_name} has been submitted by {user.Username} from {user.department.DeptName} department.</p><p>Please review and approve/reject as appropriate.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
                email_service.send_email(mgmt_user.Email, subject, html_content)
    
    flash('Consolidated MIS uploaded successfully! Management will now review it.', 'success')
    return redirect(url_for('prepare_consolidated_mis'))

@app.route('/management-consolidated-queue')
@management_required
def management_consolidated_queue():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Management is view-only, show all consolidated MIS
    consolidated_uploads = ConsolidatedMIS.query.order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    return render_template('management_consolidated_queue.html', current_user=user, consolidated_uploads=consolidated_uploads)

@app.route('/view-consolidated-mis/<int:consolidated_id>')
@management_required
def view_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    hod_uploads = MISUpload.query.filter(MISUpload.UploadID.in_([int(x) for x in consolidated.UploadedHODMISIDs.split(',')])).all() if consolidated.UploadedHODMISIDs else []
    return render_template('view_consolidated_mis.html', current_user=user, consolidated=consolidated, hod_uploads=hod_uploads)

@app.route('/approve-consolidated-mis/<int:consolidated_id>', methods=['POST'])
@management_required
def approve_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    consolidated.Status = 'Approved'
    consolidated.ApprovedDate = datetime.now(IST)
    consolidated.ApprovedBy = user.UserID
    
    hod_ids = set()
    for upload_id in consolidated.UploadedHODMISIDs.split(','):
        upload = MISUpload.query.get(int(upload_id))
        if upload:
            upload.Status = 'Approved'
            hod_ids.add(upload.UploadedBy)
    
    db.session.commit()
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[consolidated.MonthID]
    
    # Create notification for supervisor
    create_notification(consolidated.SupervisorID, 'Consolidated MIS Approved', f'Your consolidated MIS for {month_name} has been approved by Management.', 'management_approval', consolidated_id=consolidated_id)
    
    # Create notifications for all HODs in the consolidated upload
    for hod_id in hod_ids:
        create_notification(hod_id, 'MIS Approved by Management', f'Your MIS included in the consolidated MIS for {month_name} has been approved by Management.', 'management_approval')
    
    if email_service.is_configured():
        supervisor = consolidated.supervisor
        subject = f"Consolidated MIS Approved - {month_name}"
        html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial; background-color: #f9fafb;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #10b981; color: white; padding: 20px; border-radius: 5px 5px 0 0;"><h2>✓ Consolidated MIS Approved</h2></div><div style="background-color: white; padding: 30px; border: 1px solid #e5e7eb;"><p>Dear {supervisor.Username},</p><p>Your consolidated MIS for {month_name} has been <strong>approved by Management</strong>.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
        email_service.send_email(supervisor.Email, subject, html_content)
        
        # Send emails to all HODs
        for hod_id in hod_ids:
            hod = User.query.get(hod_id)
            if hod:
                subject = f"MIS Approved by Management - {month_name}"
                html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #10b981; color: white; padding: 20px; border-radius: 5px 5px 0 0;"><h2>MIS Approved</h2></div><div style="background-color: white; padding: 30px;"><p>Dear {hod.Username},</p><p>Your MIS for {month_name} has been approved by Management.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
                email_service.send_email(hod.Email, subject, html_content)
    
    flash('Consolidated MIS approved! All HOD uploads marked as approved.', 'success')
    return redirect(url_for('management_consolidated_queue'))

@app.route('/reject-consolidated-mis/<int:consolidated_id>', methods=['POST'])
@management_required
def reject_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    consolidated.Status = 'Rejected'
    db.session.commit()
    
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    month_name = month_names[consolidated.MonthID]
    
    # Get HOD IDs from consolidated uploads
    hod_ids = set()
    for upload_id in consolidated.UploadedHODMISIDs.split(','):
        upload = MISUpload.query.get(int(upload_id))
        if upload:
            hod_ids.add(upload.UploadedBy)
    
    # Create notification for supervisor
    create_notification(consolidated.SupervisorID, 'Consolidated MIS Rejected', f'Your consolidated MIS for {month_name} has been rejected by Management. Please review and resubmit.', 'management_rejection', consolidated_id=consolidated_id)
    
    # Create notifications for all HODs
    for hod_id in hod_ids:
        create_notification(hod_id, 'MIS Rejected by Management', f'Your MIS included in the consolidated MIS for {month_name} has been rejected by Management.', 'management_rejection')
    
    if email_service.is_configured():
        supervisor = consolidated.supervisor
        subject = f"Consolidated MIS Rejected - {month_name}"
        html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #ef4444; color: white; padding: 20px; border-radius: 5px 5px 0 0;"><h2>✗ Consolidated MIS Rejected</h2></div><div style="background-color: white; padding: 30px;"><p>Dear {supervisor.Username},</p><p>Your consolidated MIS for {month_name} has been rejected by Management. Please review and resubmit.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
        email_service.send_email(supervisor.Email, subject, html_content)
        
        # Send emails to all HODs
        for hod_id in hod_ids:
            hod = User.query.get(hod_id)
            if hod:
                subject = f"MIS Rejected by Management - {month_name}"
                html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #ef4444; color: white; padding: 20px; border-radius: 5px 5px 0 0;"><h2>MIS Rejected</h2></div><div style="background-color: white; padding: 30px;"><p>Dear {hod.Username},</p><p>Your MIS for {month_name} has been rejected by Management. Please review and resubmit.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
                email_service.send_email(hod.Email, subject, html_content)
    
    flash('Consolidated MIS rejected. Notifications sent to supervisor and HODs.', 'warning')
    return redirect(url_for('management_consolidated_queue'))

@app.route('/download-consolidated-mis/<int:consolidated_id>')
@login_required
def download_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    
    # Allow Admin, Management, and Supervisor to download
    if user.role.RoleName not in ['Admin', 'Management', 'Supervisor']:
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    
    from flask import send_file
    try:
        filename = consolidated.ConsolidatedFilePath.split('/')[-1]
        return send_file(consolidated.ConsolidatedFilePath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('dashboard'))

@app.route('/download-consolidated-dashboard-excel')
@management_required
def download_consolidated_dashboard_excel():
    user = User.query.get(session['user_id'])
    
    # Get filter parameters
    fy_id = request.args.get('fy_id', '')
    month_id = request.args.get('month_id', '')
    status = request.args.get('status', '')
    
    # Base query
    query = ConsolidatedMIS.query
    
    if fy_id:
        query = query.filter_by(FYID=int(fy_id))
    if month_id:
        query = query.filter_by(MonthID=int(month_id))
    if status:
        query = query.filter_by(Status=status)
    
    consolidated_reports = query.order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    
    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidated MIS Dashboard"
        
        # Headers
        headers = ['ID', 'Month', 'Financial Year', 'Supervisor', 'Department', 'Created Date', 'Status', 'Approved By', 'Approved Date']
        ws.append(headers)
        
        # Data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        for report in consolidated_reports:
            approver = User.query.get(report.ApprovedBy) if report.ApprovedBy else None
            ws.append([
                f"#{report.ConsolidatedMISID}",
                month_names[report.MonthID],
                report.financial_year.FYName,
                report.supervisor.Username or report.supervisor.EmpID,
                report.supervisor.department.DeptName,
                report.CreatedDate.strftime('%d %b %Y'),
                report.Status,
                approver.Username if approver else '',
                report.ApprovedDate.strftime('%d %b %Y') if report.ApprovedDate else ''
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Consolidated_MIS_Dashboard.xlsx')
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('consolidated_mis_dashboard'))

@app.route('/download-individual-dashboard-excel')
@management_required
def download_individual_dashboard_excel():
    user = User.query.get(session['user_id'])
    
    # Get filter parameters
    fy_id = request.args.get('fy_id', '')
    month_id = request.args.get('month_id', '')
    status = request.args.get('status', '')
    
    # Base query
    query = MISUpload.query.filter_by(IsCancelled=False)
    
    if fy_id:
        query = query.filter_by(FYID=int(fy_id))
    if month_id:
        query = query.filter_by(MonthID=int(month_id))
    if status:
        query = query.filter_by(Status=status)
    
    individual_reports = query.order_by(MISUpload.UploadDate.desc()).all()
    
    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Individual MIS Dashboard"
        
        # Headers
        headers = ['MIS Code', 'Department', 'Month', 'Financial Year', 'Uploaded By', 'Upload Date', 'Status', 'Supervisor Approved']
        ws.append(headers)
        
        # Data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        for upload in individual_reports:
            ws.append([
                upload.UploadCode,
                upload.department.DeptName,
                month_names[upload.MonthID],
                upload.financial_year.FYName,
                upload.uploader.Username or upload.uploader.EmpID,
                upload.UploadDate.strftime('%d %b %Y'),
                upload.Status,
                'Yes' if upload.SupervisorApproved else 'No'
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Individual_MIS_Dashboard.xlsx')
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('management_consolidated_reports'))

@app.route('/download-reports-excel')
@login_required
def download_reports_excel():
    user = User.query.get(session['user_id'])
    
    # Filter uploads based on role
    if user.role.RoleName == 'Admin':
        query = MISUpload.query
    elif user.role.RoleName == 'Management':
        query = MISUpload.query
    elif user.role.RoleName == 'Supervisor':
        query = MISUpload.query
    else:
        query = MISUpload.query.filter_by(DepartmentID=user.DepartmentID)
    
    query = query.filter_by(IsCancelled=False)
    
    # Apply filters
    department_id = request.args.get('department', '')
    fy_id = request.args.get('fy', '')
    status = request.args.get('status', '')
    search_code = request.args.get('search_code', '').strip()
    
    if search_code:
        query = query.filter_by(UploadCode=search_code)
    else:
        if department_id:
            query = query.filter_by(DepartmentID=int(department_id))
        if fy_id:
            query = query.filter_by(FYID=int(fy_id))
        if status:
            query = query.filter_by(Status=status)
    
    uploads = query.order_by(MISUpload.UploadDate.desc()).all()
    
    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MIS Reports"
        
        # Headers
        headers = ['MIS Code', 'Upload ID', 'Department', 'Month', 'Financial Year', 'Uploaded By', 'Upload Date', 'File Check', 'Status']
        ws.append(headers)
        
        # Data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        for upload in uploads:
            ws.append([
                upload.UploadCode,
                f"#{upload.UploadID}",
                upload.department.DeptName,
                month_names[upload.MonthID],
                upload.financial_year.FYName,
                upload.uploader.Username or upload.uploader.EmpID,
                upload.UploadDate.strftime('%d %b %Y %H:%M'),
                upload.FileCheck,
                upload.Status
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='MIS_Reports.xlsx')
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/download-my-uploads-excel')
@login_required
def download_my_uploads_excel():
    user = User.query.get(session['user_id'])
    uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).all()
    
    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "My Uploads"
        
        # Headers
        headers = ['MIS Code', 'Month', 'Financial Year', 'Uploaded By', 'Upload Date', 'File Check', 'Status']
        ws.append(headers)
        
        # Data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        for upload in uploads:
            ws.append([
                upload.UploadCode,
                month_names[upload.MonthID],
                upload.financial_year.FYName,
                upload.uploader.Username or upload.uploader.EmpID,
                upload.UploadDate.strftime('%d %b %Y %H:%M'),
                upload.FileCheck,
                'Cancelled' if upload.IsCancelled else upload.Status
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'My_Uploads_{user.department.DeptName}.xlsx')
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('my_uploads'))

@app.route('/download-approved-mis-excel')
@admin_required
def download_approved_mis_excel():
    approved_uploads = MISUpload.query.filter_by(Status='Approved').order_by(MISUpload.UploadDate.desc()).all()
    
    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Approved MIS"
        
        # Headers
        headers = ['MIS Code', 'Department', 'Month', 'Financial Year', 'Uploaded By', 'Upload Date', 'File Check', 'Status']
        ws.append(headers)
        
        # Data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        for upload in approved_uploads:
            ws.append([
                upload.UploadCode,
                upload.department.DeptName,
                month_names[upload.MonthID],
                upload.financial_year.FYName,
                upload.uploader.Username or upload.uploader.EmpID,
                upload.UploadDate.strftime('%d %b %Y %H:%M'),
                upload.FileCheck,
                upload.Status
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Approved_MIS_Reports.xlsx')
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('approved_mis'))

@app.route('/download-supervisor-uploads-excel')
@supervisor_required
def download_supervisor_uploads_excel():
    pending_uploads = MISUpload.query.filter_by(SupervisorApproved=False, Status='In Review', IsCancelled=False).order_by(MISUpload.UploadDate.desc()).all()
    
    try:
        import openpyxl
        from io import BytesIO
        from flask import send_file
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pending HOD Uploads"
        
        # Headers
        headers = ['MIS Code', 'Department', 'Month', 'Financial Year', 'Uploaded By', 'Upload Date', 'Status']
        ws.append(headers)
        
        # Data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        for upload in pending_uploads:
            ws.append([
                upload.UploadCode,
                upload.department.DeptName,
                month_names[upload.MonthID],
                upload.financial_year.FYName,
                upload.uploader.Username or upload.uploader.EmpID,
                upload.UploadDate.strftime('%d %b %Y %H:%M'),
                upload.Status
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Pending_HOD_Uploads.xlsx')
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('supervisor_uploads'))

@app.route('/admin-consolidated-management')
@admin_required
def admin_consolidated_management():
    user = User.query.get(session['user_id'])
    
    # Get filter parameters
    fy_id = request.args.get('fy_id', '')
    month_id = request.args.get('month_id', '')
    status = request.args.get('status', '')
    
    # Admin can see ALL consolidated MIS records (no restrictions)
    query = ConsolidatedMIS.query
    
    # Apply filters only if specified
    if fy_id:
        query = query.filter_by(FYID=int(fy_id))
    if month_id:
        query = query.filter_by(MonthID=int(month_id))
    if status:
        query = query.filter_by(Status=status)
    
    # Get all consolidated reports sorted by most recent first
    consolidated_reports = query.order_by(ConsolidatedMIS.CreatedDate.desc()).all()
    financial_years = FinancialYear.query.all()
    
    # Calculate statistics
    stats = {
        'total_consolidated': ConsolidatedMIS.query.count(),
        'pending_count': ConsolidatedMIS.query.filter_by(Status='Pending Review').count(),
        'approved_count': ConsolidatedMIS.query.filter_by(Status='Approved').count(),
        'rejected_count': ConsolidatedMIS.query.filter_by(Status='Rejected').count()
    }
    
    return render_template('admin_consolidated_management.html',
                         current_user=user,
                         consolidated_reports=consolidated_reports,
                         financial_years=financial_years,
                         stats=stats,
                         selected_fy=fy_id,
                         selected_month=month_id,
                         selected_status=status)

@app.route('/view-admin-consolidated-mis/<int:consolidated_id>')
@admin_required
def view_admin_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    
    hod_uploads = MISUpload.query.filter(MISUpload.UploadID.in_([int(x) for x in consolidated.UploadedHODMISIDs.split(',')])).all() if consolidated.UploadedHODMISIDs else []
    
    # Get approver info
    approver = User.query.get(consolidated.ApprovedBy) if consolidated.ApprovedBy else None
    
    return render_template('view_consolidated_mis.html',
                         current_user=user,
                         consolidated=consolidated,
                         hod_uploads=hod_uploads,
                         approver=approver)

@app.route('/edit-consolidated-mis/<int:consolidated_id>', methods=['GET', 'POST'])
@admin_required
def edit_consolidated_mis(consolidated_id):
    user = User.query.get(session['user_id'])
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    
    if request.method == 'GET':
        # Get approver info if exists
        approver = User.query.get(consolidated.ApprovedBy) if consolidated.ApprovedBy else None
        
        return render_template('edit_consolidated_mis.html',
                             current_user=user,
                             consolidated=consolidated,
                             approver=approver)
    
    # POST request - handle file and status update
    status = request.form.get('status')
    
    # Update status
    if status and status in ['Pending Review', 'Approved', 'Rejected']:
        old_status = consolidated.Status
        consolidated.Status = status
        
        # If status changed to Approved, set approval details
        if status == 'Approved' and old_status != 'Approved':
            consolidated.ApprovedBy = user.UserID
            consolidated.ApprovedDate = datetime.now(IST)
            
            # Update all included HOD uploads to Approved
            if consolidated.UploadedHODMISIDs:
                hod_upload_ids = [int(x) for x in consolidated.UploadedHODMISIDs.split(',')]
                for upload_id in hod_upload_ids:
                    upload = MISUpload.query.get(upload_id)
                    if upload:
                        upload.Status = 'Approved'
    
    # Handle file replacement
    if 'file' in request.files:
        file = request.files['file']
        
        if file and file.filename:
            if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
                flash('Only .xls or .xlsx files are allowed.', 'error')
                return redirect(url_for('edit_consolidated_mis', consolidated_id=consolidated_id))
            
            # Validate Excel file
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(temp_path)
            
            is_valid, validation_message = validate_excel_file(temp_path)
            
            if not is_valid:
                os.remove(temp_path)
                flash(f'Validation Error: {validation_message}', 'error')
                return redirect(url_for('edit_consolidated_mis', consolidated_id=consolidated_id))
            
            # Delete old file
            try:
                if os.path.exists(consolidated.ConsolidatedFilePath):
                    os.remove(consolidated.ConsolidatedFilePath)
            except Exception as e:
                flash(f'Warning: Error deleting old file: {str(e)}', 'warning')
            
            # Move new file to upload location
            fy = consolidated.financial_year
            month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'ConsolidatedMIS', fy.FYName)
            os.makedirs(upload_dir, exist_ok=True)
            
            filename = secure_filename(f"ConsolidatedMIS_{consolidated.MonthID:02d}_{fy.FYName}_{file.filename}")
            filepath = os.path.join(upload_dir, filename)
            
            os.rename(temp_path, filepath)
            
            # Update consolidated record
            consolidated.ConsolidatedFilePath = filepath
            consolidated.CreatedDate = datetime.now(IST)
    
    db.session.commit()
    flash('✓ Consolidated MIS updated successfully!', 'success')
    return redirect(url_for('admin_consolidated_management'))

@app.route('/delete-consolidated-mis/<int:consolidated_id>', methods=['POST'])
@admin_required
def delete_consolidated_mis(consolidated_id):
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    
    # Delete file from storage
    try:
        if os.path.exists(consolidated.ConsolidatedFilePath):
            os.remove(consolidated.ConsolidatedFilePath)
    except Exception as e:
        flash(f'Warning: Error deleting file: {str(e)}', 'warning')
    
    db.session.delete(consolidated)
    db.session.commit()
    
    flash('Consolidated MIS deleted successfully!', 'success')
    return redirect(url_for('admin_consolidated_management'))

def strip_html_tags(text):
    """Remove HTML tags from text"""
    import re
    if not text:
        return ''
    # Remove HTML tags
    clean = re.sub('<.*?>', '', str(text))
    return clean

@app.route('/download-consolidated-pdf/<int:consolidated_id>')
@login_required
def download_consolidated_pdf(consolidated_id):
    user = User.query.get(session['user_id'])
    
    # Allow Admin, Management, and Supervisor to download PDF
    if user.role.RoleName not in ['Admin', 'Management', 'Supervisor']:
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    consolidated = ConsolidatedMIS.query.get_or_404(consolidated_id)
    
    try:
        import openpyxl
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO
        from flask import send_file
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(consolidated.ConsolidatedFilePath)
        sheet = workbook.active
        
        # Prepare data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[consolidated.MonthID]
        
        # Get included HOD uploads
        hod_uploads = MISUpload.query.filter(MISUpload.UploadID.in_([int(x) for x in consolidated.UploadedHODMISIDs.split(',')])).all() if consolidated.UploadedHODMISIDs else []
        
        # Get approver info
        approver = User.query.get(consolidated.ApprovedBy) if consolidated.ApprovedBy else None
        
        # Extract headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        # Extract rows
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell) if cell is not None else '' for cell in row])
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#6b7280'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            borderColor=colors.HexColor('#3b82f6'),
            borderWidth=2,
            borderPadding=8,
            backColor=colors.HexColor('#eff6ff')
        )
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#374151'),
            spaceAfter=6
        )
        
        # Header
        elements.append(Paragraph("CONSOLIDATED MIS REPORT", title_style))
        elements.append(Paragraph(f"{month_name} {consolidated.financial_year.FYName}", subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Report Information Section
        elements.append(Paragraph("Report Information", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        info_data = [
            ['Report ID:', f'#{consolidated.ConsolidatedMISID}', 'Month:', month_name],
            ['Financial Year:', consolidated.financial_year.FYName, 'Status:', consolidated.Status],
            ['Created Date:', consolidated.CreatedDate.strftime('%d %b %Y, %I:%M %p'), 'Total Departments:', str(len(hod_uploads))]
        ]
        
        info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Supervisor Information Section
        elements.append(Paragraph("Prepared By", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        supervisor_data = [
            ['Name:', consolidated.supervisor.Username or consolidated.supervisor.EmpID, 'Employee ID:', consolidated.supervisor.EmpID],
            ['Department:', consolidated.supervisor.department.DeptName, 'Email:', consolidated.supervisor.Email],
            ['Role:', consolidated.supervisor.role.RoleName, 'Contact:', 'N/A']
        ]
        
        supervisor_table = Table(supervisor_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        supervisor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#93c5fd')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(supervisor_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Approval Information (if approved)
        if consolidated.Status == 'Approved' and approver:
            elements.append(Paragraph("Approved By", heading_style))
            elements.append(Spacer(1, 0.1*inch))
            
            approval_data = [
                ['Approver:', approver.Username or approver.EmpID, 'Approved Date:', consolidated.ApprovedDate.strftime('%d %b %Y, %I:%M %p') if consolidated.ApprovedDate else 'N/A'],
                ['Department:', approver.department.DeptName, 'Role:', approver.role.RoleName]
            ]
            
            approval_table = Table(approval_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
            approval_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d1fae5')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#065f46')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#6ee7b7')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(approval_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Included Departments Section
        elements.append(Paragraph(f"Included Department MIS Reports ({len(hod_uploads)})", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        dept_headers = [['No.', 'MIS Code', 'Department', 'Uploaded By', 'Upload Date', 'Status']]
        dept_data = []
        for idx, upload in enumerate(hod_uploads, 1):
            dept_data.append([
                str(idx),
                upload.UploadCode,
                upload.department.DeptName,
                upload.uploader.Username or upload.uploader.EmpID,
                upload.UploadDate.strftime('%d %b %Y'),
                'Approved' if upload.SupervisorApproved else 'Pending'
            ])
        
        dept_table = Table(dept_headers + dept_data, colWidths=[0.4*inch, 1.2*inch, 1.5*inch, 1.5*inch, 1.2*inch, 1*inch])
        dept_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#faf5ff')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#374151')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c4b5fd')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(dept_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # MIS Data Section
        elements.append(PageBreak())
        elements.append(Paragraph("Consolidated MIS Data", heading_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # Table data with all columns
        if headers and rows:
            # Calculate dynamic column widths based on content
            available_width = 10.5 * inch  # landscape A4 width minus margins
            col_count = len(headers)
            col_width = available_width / col_count if col_count > 0 else 1*inch
            
            table_data = [headers] + rows
            
            # Create table with dynamic width
            data_table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            elements.append(data_table)
        
        # Footer
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Generated on {datetime.now(IST).strftime('%d %b %Y at %I:%M %p IST')} | Confidential Report", footer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'Consolidated_MIS_{month_name}_{consolidated.financial_year.FYName}.pdf')
            
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('management_consolidated_reports'))

@app.route('/download-upload/<int:upload_id>')
@login_required
def download_upload(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions - Admin, Management, and HOD can download
    if user.role.RoleName not in ['Admin', 'HOD', 'Management', 'Supervisor'] and upload.DepartmentID != user.DepartmentID:
        flash('Access denied.', 'error')
        return redirect(url_for('reports'))
    
    from flask import send_file
    try:
        filename = upload.FilePath.split('/')[-1]
        return send_file(upload.FilePath, as_attachment=True, download_name=filename)
    except Exception as e:
        flash(f'Error downloading file: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/download-upload-pdf/<int:upload_id>')
@login_required
def download_upload_pdf(upload_id):
    user = User.query.get(session['user_id'])
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions - Admin, Management, and HOD can download
    if user.role.RoleName not in ['Admin', 'HOD', 'Management', 'Supervisor'] and upload.DepartmentID != user.DepartmentID:
        flash('Access denied.', 'error')
        return redirect(url_for('reports'))
    
    try:
        import openpyxl
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO
        from flask import send_file
        
        # Load the Excel file
        workbook = openpyxl.load_workbook(upload.FilePath)
        sheet = workbook.active
        
        # Prepare data
        month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        month_name = month_names[upload.MonthID]
        
        # Get supervisor approval info
        supervisor_approver = User.query.get(upload.SupervisorApprovedBy) if upload.SupervisorApprovedBy else None
        
        # Extract headers
        headers = []
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value))
        
        # Extract rows
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(cell) if cell is not None else '' for cell in row])
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=30)
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#6b7280'),
            spaceAfter=20,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold',
            borderColor=colors.HexColor('#3b82f6'),
            borderWidth=2,
            borderPadding=8,
            backColor=colors.HexColor('#eff6ff')
        )
        
        # Header
        elements.append(Paragraph("DEPARTMENTAL MIS REPORT", title_style))
        elements.append(Paragraph(f"{upload.department.DeptName} Department", subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Report Information Section
        elements.append(Paragraph("Report Information", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        info_data = [
            ['MIS Code:', upload.UploadCode, 'Upload ID:', f'#{upload.UploadID}'],
            ['Month:', month_name, 'Financial Year:', upload.financial_year.FYName],
            ['Department:', upload.department.DeptName, 'Upload Date:', upload.UploadDate.strftime('%d %b %Y, %I:%M %p')],
            ['File Check:', upload.FileCheck, 'Status:', upload.Status]
        ]
        
        info_table = Table(info_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#374151')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Uploader Information Section
        elements.append(Paragraph("Uploaded By", heading_style))
        elements.append(Spacer(1, 0.1*inch))
        
        uploader_data = [
            ['Name:', upload.uploader.Username or upload.uploader.EmpID, 'Employee ID:', upload.uploader.EmpID],
            ['Department:', upload.uploader.department.DeptName, 'Email:', upload.uploader.Email],
            ['Role:', upload.uploader.role.RoleName, 'Upload Date:', upload.UploadDate.strftime('%d %b %Y, %I:%M %p')]
        ]
        
        uploader_table = Table(uploader_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        uploader_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#93c5fd')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(uploader_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Approval Information
        if upload.SupervisorApproved and supervisor_approver:
            elements.append(Paragraph("Supervisor Approval", heading_style))
            elements.append(Spacer(1, 0.1*inch))
            
            approval_data = [
                ['Approved By:', supervisor_approver.Username or supervisor_approver.EmpID, 'Approval Date:', upload.SupervisorApprovedDate.strftime('%d %b %Y, %I:%M %p') if upload.SupervisorApprovedDate else 'N/A'],
                ['Supervisor Dept:', supervisor_approver.department.DeptName, 'Status:', 'Approved']
            ]
            
            approval_table = Table(approval_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
            approval_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d1fae5')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#065f46')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#6ee7b7')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(approval_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Management Approval (if status is Approved)
        if upload.Status == 'Approved':
            elements.append(Paragraph("Management Approval", heading_style))
            elements.append(Spacer(1, 0.1*inch))
            
            mgmt_data = [
                ['Status:', 'Approved by Management', 'Final Status:', upload.Status]
            ]
            
            mgmt_table = Table(mgmt_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
            mgmt_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#d1fae5')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#065f46')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#6ee7b7')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(mgmt_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # Additional Metadata
        if upload.IsModified or upload.IsCancelled:
            elements.append(Paragraph("Additional Information", heading_style))
            elements.append(Spacer(1, 0.1*inch))
            
            metadata = []
            if upload.IsModified:
                metadata.append(['Modified:', 'Yes - This report has been modified after initial submission'])
            if upload.IsCancelled:
                metadata.append(['Cancelled:', 'Yes - This report was cancelled by the uploader'])
            
            metadata_table = Table(metadata, colWidths=[1.5*inch, 5.5*inch])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#fef3c7')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#78350f')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('PADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#fde68a')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(metadata_table)
            elements.append(Spacer(1, 0.2*inch))
        
        # MIS Data Section
        elements.append(PageBreak())
        elements.append(Paragraph("Departmental MIS Data", heading_style))
        elements.append(Spacer(1, 0.15*inch))
        
        # Table data with all columns
        if headers and rows:
            # Calculate dynamic column widths
            available_width = 10.5 * inch
            col_count = len(headers)
            col_width = available_width / col_count if col_count > 0 else 1*inch
            
            table_data = [headers] + rows
            
            data_table = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            data_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f9fafb')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
            ]))
            elements.append(data_table)
        
        # Footer
        elements.append(Spacer(1, 0.3*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#6b7280'),
            alignment=TA_CENTER
        )
        elements.append(Paragraph(f"Generated on {datetime.now(IST).strftime('%d %b %Y at %I:%M %p IST')} | MIS Code: {upload.UploadCode} | Confidential Report", footer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f'MIS_{upload.UploadCode}_{upload.department.DeptName}_{month_name}_{upload.financial_year.FYName}.pdf')
            
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/delete-upload/<int:upload_id>', methods=['POST'])
@login_required
def delete_upload(upload_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Admin can delete any upload (In Review or Approved, any time)
    if user.role.RoleName == 'Admin':
        try:
            if os.path.exists(upload.FilePath):
                os.remove(upload.FilePath)
        except Exception as e:
            flash(f'Warning: File deletion error: {str(e)}', 'warning')
        
        db.session.delete(upload)
        db.session.commit()
        flash('Upload deleted successfully!', 'success')
        return redirect(url_for('approved_mis') if upload.Status == 'Approved' else url_for('reports'))
    
    # HOD can only delete "In Review" and "Rejected" uploads from their department
    if user.role.RoleName == 'HOD':
        if upload.Status not in ['In Review', 'Rejected']:
            flash(f'Cannot delete uploads with status "{upload.Status}". Only "In Review" and "Rejected" uploads can be deleted.', 'error')
            return redirect(url_for('my_uploads'))
        
        if upload.DepartmentID != user.DepartmentID:
            flash('You can only delete uploads from your department.', 'error')
            return redirect(url_for('my_uploads'))
        
        try:
            if os.path.exists(upload.FilePath):
                os.remove(upload.FilePath)
        except Exception as e:
            flash(f'Warning: File deletion error: {str(e)}', 'warning')
        
        # Mark as cancelled instead of deleting so Management can see it was cancelled
        upload.IsCancelled = True
        db.session.commit()
        flash('Upload marked as cancelled. Management will see this change.', 'success')
        return redirect(url_for('my_uploads'))
    
    # Other roles cannot delete
    flash('You do not have permission to delete uploads.', 'error')
    return redirect(url_for('reports'))

@app.route('/edit-upload/<int:upload_id>', methods=['GET', 'POST'])
@login_required
def edit_upload(upload_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    upload = MISUpload.query.get_or_404(upload_id)
    
    # Check permissions
    # Admin can edit any upload anytime
    # HOD can only edit "In Review" uploads from their department
    if user.role.RoleName == 'Admin':
        pass  # Admin can edit any upload
    elif user.role.RoleName == 'HOD':
        if upload.DepartmentID != user.DepartmentID:
            flash('You can only edit uploads from your department.', 'error')
            return redirect(url_for('my_uploads'))
        
        if upload.Status not in ['In Review', 'Rejected']:
            flash(f'Cannot modify uploads with status "{upload.Status}". Only "In Review" and "Rejected" uploads can be modified.', 'error')
            return redirect(url_for('my_uploads'))
    else:
        flash('You do not have permission to edit uploads.', 'error')
        return redirect(url_for('reports'))
    
    # GET request - show edit form
    if request.method == 'GET':
        financial_years = FinancialYear.query.all()
        departments = Department.query.all() if user.role.RoleName == 'Admin' else Department.query.filter_by(DeptID=user.DepartmentID).all()
        
        return render_template('edit_upload.html',
                             current_user=user,
                             upload=upload,
                             departments=departments,
                             financial_years=financial_years)
    
    # POST request - handle file update
    if 'file' in request.files:
        file = request.files['file']
        
        if file and file.filename:
            if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
                flash('Only .xls or .xlsx files are allowed.', 'error')
                return redirect(url_for('edit_upload', upload_id=upload_id))
            
            # Validate Excel file
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(file.filename))
            file.save(temp_path)
            
            is_valid, validation_message = validate_excel_file(temp_path)
            
            if not is_valid:
                os.remove(temp_path)
                flash(f'Validation Error: {validation_message}', 'error')
                return redirect(url_for('edit_upload', upload_id=upload_id))
            
            # Delete old file
            try:
                if os.path.exists(upload.FilePath):
                    os.remove(upload.FilePath)
            except Exception as e:
                flash(f'Warning: Error deleting old file: {str(e)}', 'warning')
            
            # Move new file to upload location
            fy = upload.financial_year
            dept = upload.department
            month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
            
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], fy.FYName, dept.DeptName, month_names[upload.MonthID])
            os.makedirs(upload_dir, exist_ok=True)
            
            filename = secure_filename(f"{dept.DeptName}_{month_names[upload.MonthID]}_{fy.FYName}_{file.filename}")
            filepath = os.path.join(upload_dir, filename)
            
            os.rename(temp_path, filepath)
            
            # Update upload record
            upload.FilePath = filepath
            upload.FileCheck = 'Validated'
            upload.UploadDate = datetime.now(IST)
            # Mark as modified for HOD so Management can see changes
            if user.role.RoleName == 'HOD':
                upload.IsModified = True
                # Reset status to In Review if it was Rejected
                if upload.Status == 'Rejected':
                    upload.Status = 'In Review'
                    upload.SupervisorApproved = False
                    upload.SupervisorApprovedBy = None
                    upload.SupervisorApprovedDate = None
            db.session.commit()
            
            flash('✓ File updated successfully! Management will see this upload was modified.', 'success')
            return redirect(url_for('my_uploads'))
    
    flash('No file provided for update.', 'error')
    return redirect(url_for('edit_upload', upload_id=upload_id))

@app.route('/config-master')
@admin_required
def config_master():
    user = User.query.get(session['user_id'])
    companies = Company.query.all()
    departments = Department.query.all()
    financial_years = FinancialYear.query.all()
    
    return render_template('config_master.html', 
                                 current_user=user,
                                 companies=companies,
                                 departments=departments,
                                 financial_years=financial_years)

@app.route('/department-management')
@admin_required
def department_management():
    user = User.query.get(session['user_id'])
    departments = Department.query.all()
    
    return render_template('department_management.html',
                                 current_user=user,
                                 departments=departments)

@app.route('/add-company', methods=['POST'])
@admin_required
def add_company():
    company_name = request.form.get('company_name')
    
    if not company_name:
        flash('Company name is required.', 'error')
    elif Company.query.filter_by(CompanyName=company_name).first():
        flash('Company already exists.', 'error')
    else:
        company = Company(CompanyName=company_name, ActiveFlag=True)  # type: ignore
        db.session.add(company)
        db.session.commit()
        flash('Company added successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/toggle-company/<int:company_id>', methods=['POST'])
@admin_required
def toggle_company(company_id):
    company = Company.query.get_or_404(company_id)
    company.ActiveFlag = not company.ActiveFlag
    db.session.commit()
    flash(f'Company {company.CompanyName} {"activated" if company.ActiveFlag else "deactivated"}.', 'success')
    return redirect(url_for('config_master'))

@app.route('/edit-company/<int:company_id>', methods=['POST'])
@admin_required
def edit_company(company_id):
    company = Company.query.get_or_404(company_id)
    new_name = request.form.get('company_name')
    
    if not new_name:
        flash('Company name is required.', 'error')
    elif new_name != company.CompanyName and Company.query.filter_by(CompanyName=new_name).first():
        flash('Company name already exists.', 'error')
    else:
        company.CompanyName = new_name
        db.session.commit()
        flash(f'Company updated successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/delete-company/<int:company_id>', methods=['POST'])
@admin_required
def delete_company(company_id):
    company = Company.query.get_or_404(company_id)
    
    db.session.delete(company)
    db.session.commit()
    flash(f'Company {company.CompanyName} deleted successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/add-department', methods=['POST'])
@admin_required
def add_department():
    dept_name = request.form.get('dept_name')
    
    if not dept_name:
        flash('Department name is required.', 'error')
    elif Department.query.filter_by(DeptName=dept_name).first():
        flash('Department already exists.', 'error')
    else:
        department = Department(DeptName=dept_name, ActiveFlag=True)  # type: ignore
        db.session.add(department)
        db.session.commit()
        flash('Department added successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/toggle-department/<int:dept_id>', methods=['POST'])
@admin_required
def toggle_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    dept.ActiveFlag = not dept.ActiveFlag
    db.session.commit()
    flash(f'Department {dept.DeptName} {"activated" if dept.ActiveFlag else "deactivated"}.', 'success')
    return redirect(url_for('config_master'))

@app.route('/edit-department/<int:dept_id>', methods=['POST'])
@admin_required
def edit_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    new_name = request.form.get('dept_name')
    
    if not new_name:
        flash('Department name is required.', 'error')
    elif new_name != dept.DeptName and Department.query.filter_by(DeptName=new_name).first():
        flash('Department name already exists.', 'error')
    else:
        dept.DeptName = new_name
        db.session.commit()
        flash(f'Department updated successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/delete-department/<int:dept_id>', methods=['POST'])
@admin_required
def delete_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    
    # Check if department has associated users or uploads
    if dept.users or dept.uploads or dept.templates:
        flash('Cannot delete department with associated users, uploads, or templates. Deactivate it instead.', 'error')
    else:
        db.session.delete(dept)
        db.session.commit()
        flash(f'Department {dept.DeptName} deleted successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/add-fy', methods=['POST'])
@admin_required
def add_fy():
    fy_name = request.form.get('fy_name')
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not fy_name or not start_date_str or not end_date_str:
        flash('All fields are required.', 'error')
        return redirect(url_for('config_master'))
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    if FinancialYear.query.filter_by(FYName=fy_name).first():
        flash('Financial Year already exists.', 'error')
    elif start_date >= end_date:
        flash('Start date must be before end date.', 'error')
    else:
        fy = FinancialYear(FYName=fy_name, StartDate=start_date, EndDate=end_date, ActiveFlag=False)  # type: ignore
        db.session.add(fy)
        db.session.commit()
        flash('Financial Year added successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/toggle-fy/<int:fy_id>', methods=['POST'])
@admin_required
def toggle_fy(fy_id):
    fy = FinancialYear.query.get_or_404(fy_id)
    
    if not fy.ActiveFlag:
        FinancialYear.query.update({FinancialYear.ActiveFlag: False})
        fy.ActiveFlag = True
        db.session.commit()
        flash(f'Financial Year {fy.FYName} is now active.', 'success')
    else:
        flash('This Financial Year is already active.', 'error')
    
    return redirect(url_for('config_master'))

@app.route('/edit-fy/<int:fy_id>', methods=['POST'])
@admin_required
def edit_fy(fy_id):
    fy = FinancialYear.query.get_or_404(fy_id)
    fy_name = request.form.get('fy_name')
    start_date_str = request.form.get('start_date')
    end_date_str = request.form.get('end_date')
    
    if not fy_name or not start_date_str or not end_date_str:
        flash('All fields are required.', 'error')
        return redirect(url_for('config_master'))
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    if fy_name != fy.FYName and FinancialYear.query.filter_by(FYName=fy_name).first():
        flash('Financial Year name already exists.', 'error')
    elif start_date >= end_date:
        flash('Start date must be before end date.', 'error')
    else:
        fy.FYName = fy_name
        fy.StartDate = start_date
        fy.EndDate = end_date
        db.session.commit()
        flash('Financial Year updated successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/delete-fy/<int:fy_id>', methods=['POST'])
@admin_required
def delete_fy(fy_id):
    fy = FinancialYear.query.get_or_404(fy_id)
    
    if fy.uploads:
        flash('Cannot delete Financial Year with associated uploads. Deactivate it instead.', 'error')
    else:
        db.session.delete(fy)
        db.session.commit()
        flash(f'Financial Year {fy.FYName} deleted successfully!', 'success')
    
    return redirect(url_for('config_master'))

@app.route('/send-test-email', methods=['POST'])
@admin_required
def send_test_email():
    """Send a test email to verify email configuration"""
    user = User.query.get(session['user_id'])
    
    if not email_service.is_configured():
        flash('Email service is not configured. Please set up SMTP credentials in email_config.py.', 'error')
        return redirect(url_for('dashboard'))
    
    # Send test email to the admin user
    subject = "Test Email - MIS System"
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
            .header {{ background-color: #4F46E5; color: white; padding: 20px; border-radius: 5px 5px 0 0; }}
            .content {{ background-color: #f9fafb; padding: 30px; border: 1px solid #e5e7eb; }}
            .footer {{ background-color: #f3f4f6; padding: 15px; text-align: center; border-radius: 0 0 5px 5px; font-size: 12px; color: #6b7280; }}
            .success-box {{ background-color: #d1fae5; border-left: 4px solid #10b981; padding: 15px; margin: 20px 0; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h2 style="margin: 0;">✓ Email Test Successful</h2>
            </div>
            <div class="content">
                <div class="success-box">
                    <strong>Congratulations!</strong> Your email configuration is working correctly.
                </div>
                
                <p>Dear {user.Username},</p>
                
                <p>This is a test email from the MIS Configuration System to verify that your SMTP settings are configured correctly.</p>
                
                <h3>Configuration Details:</h3>
                <ul>
                    <li><strong>SMTP Host:</strong> {email_service.smtp_host}</li>
                    <li><strong>SMTP Port:</strong> {email_service.smtp_port}</li>
                    <li><strong>From Email:</strong> {email_service.from_email}</li>
                    <li><strong>Sent to:</strong> {user.Email}</li>
                    <li><strong>Timestamp:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</li>
                </ul>
                
                <p>If you received this email, your email notification system is ready to use!</p>
                
                <p>Best regards,<br>
                <strong>MIS System</strong></p>
            </div>
            <div class="footer">
                This is a test email from the MIS Upload System.
            </div>
        </div>
    </body>
    </html>
    """
    
    text_content = f"""
Test Email - MIS System

✓ Email Configuration Test Successful

Dear {user.Username},

This is a test email from the MIS Configuration System to verify that your SMTP settings are configured correctly.

Configuration Details:
- SMTP Host: {email_service.smtp_host}
- SMTP Port: {email_service.smtp_port}
- From Email: {email_service.from_email}
- Sent to: {user.Email}
- Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

If you received this email, your email notification system is ready to use!

Best regards,
MIS System
    """
    
    success, message = email_service.send_email(user.Email, subject, html_content, text_content)
    
    if success:
        flash(f'✓ Test email sent successfully to {user.Email}! Check your inbox.', 'success')
    else:
        flash(f'✗ Failed to send test email: {message}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/send-upload-notifications', methods=['POST'])
@admin_required
def send_upload_notifications():
    """Send email notifications to all HOD users about MIS upload window"""
    hod_role = Role.query.filter_by(RoleName='HOD').first()
    
    if not hod_role:
        flash('HOD role not found in the system.', 'error')
        return redirect(url_for('dashboard'))
    
    hod_users = User.query.filter_by(RoleID=hod_role.RoleID, IsActive=True).all()
    
    if not hod_users:
        flash('No active HOD users found to notify.', 'warning')
        return redirect(url_for('dashboard'))
    
    if not email_service.is_configured():
        flash('Email service is not configured. Please set up SMTP credentials in environment variables.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get application URL from environment or request
    app_url = os.environ.get('REPLIT_DEV_DOMAIN')
    if app_url:
        app_url = f'https://{app_url}'
    else:
        app_url = request.host_url.rstrip('/')
    
    success_count, total_count, messages = email_service.send_upload_window_notification(hod_users, app_url)
    
    # Show detailed results
    if success_count == total_count:
        flash(f'✓ Successfully sent notifications to all {total_count} HOD users!', 'success')
    elif success_count > 0:
        flash(f'⚠ Partially successful: Sent {success_count} out of {total_count} notifications.', 'warning')
    else:
        flash(f'✗ Failed to send any notifications. Please check email configuration.', 'error')
    
    # Show individual failures
    for msg in messages:
        if 'Failed' in msg or 'error' in msg.lower() or 'Authentication' in msg:
            flash(f'❌ {msg}', 'error')
    
    return redirect(url_for('dashboard'))

@app.route('/user-management')
@admin_required
def user_management():
    user = User.query.get(session['user_id'])
    users = User.query.all()
    departments = Department.query.all()
    roles = Role.query.all()
    email_configured = email_service.is_configured()
    
    return render_template('user_management.html',
                                 current_user=user,
                                 users=users,
                                 departments=departments,
                                 roles=roles,
                                 email_configured=email_configured)

@app.route('/add-user', methods=['POST'])
@admin_required
def add_user():
    emp_id = request.form.get('emp_id')
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    department_id = request.form.get('department_id')
    role_id = request.form.get('role_id')
    
    if not emp_id or not email or not password or not department_id or not role_id:
        flash('Employee ID, Email, Password, Department and Role are required fields.', 'error')
    
    elif User.query.filter_by(EmpID=emp_id).first():
        flash('Employee ID already exists.', 'error')
    elif User.query.filter_by(Email=email).first():
        flash('Email already exists.', 'error')
    else:
        hashed_password = hash_password(password)
        new_user = User(  # type: ignore
            EmpID=emp_id,
            Username=username if username else None,
            Email=email,
            PasswordHash=hashed_password,
            DepartmentID=department_id,
            RoleID=role_id,
            IsActive=True,
            FailedLoginAttempts=0,
            PasswordLastChanged=datetime.now(IST),
            PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
        )
        db.session.add(new_user)
        db.session.commit()
        flash('User created successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/edit-user/<int:user_id>', methods=['POST'])
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    emp_id = request.form.get('emp_id')
    username = request.form.get('username')
    email = request.form.get('email')
    department_id = request.form.get('department_id')
    role_id = request.form.get('role_id')
    password = request.form.get('password')
    
    if not emp_id or not email or not department_id or not role_id:
        flash('Employee ID, Email, Department and Role are required fields.', 'error')
        return redirect(url_for('user_management'))
    
    if not email.lower().endswith('@gmail.com'):
        flash('Only Gmail addresses are allowed. Please use an email ending with @gmail.com', 'error')
    elif emp_id != user.EmpID and User.query.filter_by(EmpID=emp_id).first():
        flash('Employee ID already exists.', 'error')
    elif email != user.Email and User.query.filter_by(Email=email).first():
        flash('Email already exists.', 'error')
    else:
        user.EmpID = emp_id
        user.Username = username if username else None
        user.Email = email
        user.DepartmentID = department_id
        user.RoleID = role_id
        if password:
            user.PasswordHash = hash_password(password)
            user.PasswordLastChanged = datetime.now(IST)
            user.PasswordExpiryDate = datetime.now(IST) + timedelta(days=90)
        db.session.commit()
        flash('User updated successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/toggle-user/<int:user_id>', methods=['POST'])
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.IsActive = not user.IsActive
    db.session.commit()
    flash(f'User {user.EmpID} {"activated" if user.IsActive else "deactivated"}.', 'success')
    return redirect(url_for('user_management'))

@app.route('/delete-user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    
    if user.uploads:
        flash('Cannot delete user with associated uploads. Deactivate instead.', 'error')
    else:
        db.session.delete(user)
        db.session.commit()
        flash(f'User {user.Username} deleted successfully!', 'success')
    
    return redirect(url_for('user_management'))

@app.route('/mis-upload')
@login_required
def mis_upload():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Admin and HOD can upload during window, Management cannot upload
    if user.role.RoleName == 'Admin':
        upload_allowed, upload_message = check_upload_window()
        departments = Department.query.all()
        uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).all()
        hod_blocked_message = None
    elif user.role.RoleName == 'Management':
        upload_allowed = False
        upload_message = "Management role cannot upload MIS. Please use the Approval Queue to review and approve uploads."
        departments = Department.query.all()
        uploads = MISUpload.query.order_by(MISUpload.UploadDate.desc()).all()
        hod_blocked_message = None
    else:
        upload_allowed, upload_message = check_upload_window()
        # HOD can only see their department
        departments = Department.query.filter_by(DeptID=user.DepartmentID).all()
        uploads = MISUpload.query.filter_by(DepartmentID=user.DepartmentID).order_by(MISUpload.UploadDate.desc()).all()
        
        # Check if HOD already has an approved MIS for current month
        current_month = date.today().month
        active_fy = FinancialYear.query.filter_by(ActiveFlag=True).first()
        fy_id = active_fy.FYID if active_fy else None
        
        approved_upload = MISUpload.query.filter_by(
            DepartmentID=user.DepartmentID,
            MonthID=current_month,
            FYID=fy_id,
            Status='Approved'
        ).first() if fy_id else None
        
        if approved_upload:
            upload_allowed = False
            hod_blocked_message = f"An approved MIS upload already exists for {['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][current_month]}. No new uploads are allowed for this period."
        else:
            hod_blocked_message = None
    
    financial_years = FinancialYear.query.all()
    
    return render_template('mis_upload.html',
                                 current_user=user,
                                 upload_allowed=upload_allowed,
                                 upload_message=upload_message,
                                 hod_blocked_message=hod_blocked_message,
                                 current_date=date.today().strftime('%Y-%m-%d'),
                                 departments=departments,
                                 financial_years=financial_years,
                                 upload_window_start=UPLOAD_WINDOW_START_DAY,
                                 upload_window_end=UPLOAD_WINDOW_END_DAY,
                                 uploads=uploads)

@app.route('/upload-mis', methods=['POST'])
@login_required
def upload_mis():
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    # Management cannot upload
    if user.role.RoleName == 'Management':
        flash('Management role cannot upload MIS files.', 'error')
        return redirect(url_for('mis_upload'))
    
    # HOD must follow upload window, but Admin can upload anytime
    if user.role.RoleName != 'Admin':
        upload_allowed, upload_message = check_upload_window()
        if not upload_allowed:
            flash(upload_message, 'error')
            return redirect(url_for('mis_upload'))
    
    if 'file' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('mis_upload'))
    
    file = request.files['file']
    
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('mis_upload'))
    
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        flash('Only .xls or .xlsx files are allowed.', 'error')
        return redirect(url_for('mis_upload'))
    
    department_id = request.form.get('department_id')
    month_id = request.form.get('month_id')
    fy_id = request.form.get('fy_id')
    
    if not department_id or not month_id or not fy_id:
        flash('All fields are required.', 'error')
        return redirect(url_for('mis_upload'))
    
    # Get FY and Department names for organized storage
    fy = FinancialYear.query.get(fy_id)
    dept = Department.query.get(department_id)
    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
    
    # Check for duplicate uploads - same department, month, and FY (not cancelled)
    existing_upload = MISUpload.query.filter_by(
        DepartmentID=int(department_id),
        MonthID=int(month_id),
        FYID=int(fy_id),
        IsCancelled=False
    ).first()
    
    if existing_upload:
        flash(f'MIS upload already exists for {dept.DeptName} - {month_names[int(month_id)]} {fy.FYName}. Please delete or modify the existing upload instead.', 'error')
        return redirect(url_for('mis_upload'))
    
    # HOD users can only upload for current month
    if user.role.RoleName == 'HOD':
        current_month = date.today().month
        if int(month_id) != current_month:
            flash(f'HOD users can only upload MIS data for the current month (Month {current_month}). Old month uploads are not allowed.', 'error')
            return redirect(url_for('mis_upload'))
    
    # HOD can only upload for their own department
    if user.role.RoleName == 'HOD' and int(department_id) != user.DepartmentID:
        flash('You can only upload for your own department.', 'error')
        return redirect(url_for('mis_upload'))
    
    # Create organized directory structure: /MISUploads/[FY]/[Department]/[Month]/
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], fy.FYName, dept.DeptName, month_names[int(month_id)])
    os.makedirs(upload_dir, exist_ok=True)
    
    # Allow multiple uploads for the same month - Management will review all of them
    # No automatic archiving - all uploads with same month stay "In Review" for Management decision
    
    # Save new file
    filename = secure_filename(f"{dept.DeptName}_{month_names[int(month_id)]}_{fy.FYName}_{file.filename}")
    filepath = os.path.join(upload_dir, filename)
    file.save(filepath)
    
    # Validate Excel file content
    is_valid, validation_message = validate_excel_file(filepath)
    
    if not is_valid:
        os.remove(filepath)
        flash(f'Validation Error: {validation_message}', 'error')
        return redirect(url_for('mis_upload'))
    
    # Admin uploads are auto-approved
    if user.role.RoleName == 'Admin':
        upload_status = 'Approved'
        supervisor_approved = True
    else:
        upload_status = 'In Review'
        supervisor_approved = False
    
    # Generate unique MIS code
    mis_code = generate_mis_code(int(department_id))
    
    upload = MISUpload(  # type: ignore
        UploadCode=mis_code,
        DepartmentID=department_id,
        MonthID=month_id,
        FYID=fy_id,
        UploadedBy=session['user_id'],
        FilePath=filepath,
        FileCheck='Validated',
        Status=upload_status,
        SupervisorApproved=supervisor_approved,
        SupervisorApprovedBy=session['user_id'] if supervisor_approved else None,
        SupervisorApprovedDate=datetime.now(IST) if supervisor_approved else None
    )
    db.session.add(upload)
    db.session.commit()
    
    flash(f'✓ File Validation Success: {validation_message} File uploaded successfully and is now pending Supervisor review.', 'success')
    return redirect(url_for('mis_upload'))

@app.route('/template-management')
@admin_required
def template_management():
    user = User.query.get(session['user_id'])
    departments = Department.query.all()
    templates = Template.query.all()
    
    return render_template('template_management.html',
                                 current_user=user,
                                 departments=departments,
                                 templates=templates)

@app.route('/download-template/<int:dept_id>')
@login_required
def download_template(dept_id):
    user = User.query.get(session['user_id'])
    if not user or not user.IsActive:
        session.clear()
        flash('Your session has expired. Please log in again.', 'error')
        return redirect(url_for('login'))
    
    template = Template.query.filter_by(DepartmentID=dept_id).order_by(Template.UploadDate.desc()).first()
    
    if not template:
        flash('No template found for this department.', 'error')
        return redirect(url_for('mis_upload'))
    
    from flask import send_file
    try:
        return send_file(template.FilePath, as_attachment=True, download_name=f"MIS_Template_{dept_id}.xlsx")
    except Exception as e:
        flash(f'Error downloading template: {str(e)}', 'error')
        return redirect(url_for('mis_upload'))

@app.route('/upload-template', methods=['POST'])
@admin_required
def upload_template():
    if 'file' not in request.files:
        flash('No file selected.', 'error')
        return redirect(url_for('template_management'))
    
    file = request.files['file']
    
    if not file or not file.filename:
        flash('No file selected.', 'error')
        return redirect(url_for('template_management'))
    
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        flash('Only .xls or .xlsx files are allowed.', 'error')
        return redirect(url_for('template_management'))
    
    department_id = request.form.get('department_id')
    
    if not department_id:
        flash('Department is required.', 'error')
        return redirect(url_for('template_management'))
    
    filename = secure_filename(f"template_{department_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    template = Template(  # type: ignore
        DepartmentID=department_id,
        FilePath=filepath
    )
    db.session.add(template)
    db.session.commit()
    
    flash('Template uploaded successfully!', 'success')
    return redirect(url_for('template_management'))

@app.route('/edit-template/<int:template_id>', methods=['GET', 'POST'])
@admin_required
def edit_template(template_id):
    template = Template.query.get_or_404(template_id)
    user = User.query.get(session['user_id'])
    departments = Department.query.all()
    
    if request.method == 'POST':
        file = request.files.get('file')
        department_id = request.form.get('department_id')
        
        if not department_id:
            flash('Department is required.', 'error')
            return render_template('edit_template.html', template=template, departments=departments, current_user=user)
        
        # If new file provided, replace the old one
        if file and file.filename:
            if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
                flash('Only .xls or .xlsx files are allowed.', 'error')
                return render_template('edit_template.html', template=template, departments=departments, current_user=user)
            
            # Delete old file
            if os.path.exists(template.FilePath):
                os.remove(template.FilePath)
            
            # Save new file
            filename = secure_filename(f"template_{department_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            template.FilePath = filepath
            template.UploadDate = datetime.now(IST)
        
        template.DepartmentID = department_id
        db.session.commit()
        flash('Template updated successfully!', 'success')
        return redirect(url_for('template_management'))
    
    return render_template('edit_template.html', template=template, departments=departments, current_user=user)

@app.route('/delete-template/<int:template_id>', methods=['POST'])
@admin_required
def delete_template(template_id):
    template = Template.query.get_or_404(template_id)
    
    # Delete file from storage
    if os.path.exists(template.FilePath):
        os.remove(template.FilePath)
    
    db.session.delete(template)
    db.session.commit()
    
    flash('Template deleted successfully!', 'success')
    return redirect(url_for('template_management'))

def init_db():
    with app.app_context():
        db.create_all()
        
        if Role.query.count() == 0:
            roles = [
                Role(RoleName='Admin'),  # type: ignore
                Role(RoleName='Management'),  # type: ignore
                Role(RoleName='Supervisor'),  # type: ignore
                Role(RoleName='HOD')  # type: ignore
            ]
            db.session.add_all(roles)
            db.session.commit()
            print("Roles seeded.")
        
        if Department.query.count() == 0:
            departments = [
                Department(DeptName='Finance', ActiveFlag=True),  # type: ignore
                Department(DeptName='HR', ActiveFlag=True),  # type: ignore
                Department(DeptName='IT', ActiveFlag=True)  # type: ignore
            ]
            db.session.add_all(departments)
            db.session.commit()
            print("Departments seeded.")
        
        if Company.query.count() == 0:
            company = Company(CompanyName='Default Company', ActiveFlag=True)  # type: ignore
            db.session.add(company)
            db.session.commit()
            print("Company seeded.")
        
        if FinancialYear.query.count() == 0:
            fy = FinancialYear(  # type: ignore
                FYName='2024-2025',
                StartDate=date(2024, 4, 1),
                EndDate=date(2025, 3, 31),
                ActiveFlag=True
            )
            db.session.add(fy)
            db.session.commit()
            print("Financial Year seeded.")
        
        if User.query.count() == 0:
            admin_role = Role.query.filter_by(RoleName='Admin').first()
            management_role = Role.query.filter_by(RoleName='Management').first()
            supervisor_role = Role.query.filter_by(RoleName='Supervisor').first()
            hod_role = Role.query.filter_by(RoleName='HOD').first()
            
            finance_dept = Department.query.filter_by(DeptName='Finance').first()
            hr_dept = Department.query.filter_by(DeptName='HR').first()
            it_dept = Department.query.filter_by(DeptName='IT').first()
            
            if admin_role and management_role and supervisor_role and hod_role and finance_dept and hr_dept and it_dept:
                users = [
                    User(  # type: ignore
                        EmpID='EMP001',
                        Username='Admin User',
                        Email='admin@gmail.com',
                        PasswordHash=hash_password('admin123'),
                        DepartmentID=it_dept.DeptID,
                        RoleID=admin_role.RoleID,
                        IsActive=True,
                        FailedLoginAttempts=0,
                        PasswordLastChanged=datetime.now(IST),
                        PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
                    ),
                    User(  # type: ignore
                        EmpID='EMP005',
                        Username='Management',
                        Email='management@gmail.com',
                        PasswordHash=hash_password('manager123'),
                        DepartmentID=finance_dept.DeptID,
                        RoleID=management_role.RoleID,
                        IsActive=True,
                        FailedLoginAttempts=0,
                        PasswordLastChanged=datetime.now(IST),
                        PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
                    ),
                    User(  # type: ignore
                        EmpID='EMP006',
                        Username='Supervisor',
                        Email='supervisor@gmail.com',
                        PasswordHash=hash_password('supervisor123'),
                        DepartmentID=finance_dept.DeptID,
                        RoleID=supervisor_role.RoleID,
                        IsActive=True,
                        FailedLoginAttempts=0,
                        PasswordLastChanged=datetime.now(IST),
                        PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
                    ),
                    User(  # type: ignore
                        EmpID='EMP002',
                        Username='HOD Finance',
                        Email='hod.finance@gmail.com',
                        PasswordHash=hash_password('hod123'),
                        DepartmentID=finance_dept.DeptID,
                        RoleID=hod_role.RoleID,
                        IsActive=True,
                        FailedLoginAttempts=0,
                        PasswordLastChanged=datetime.now(IST),
                        PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
                    ),
                    User(  # type: ignore
                        EmpID='EMP003',
                        Username='HOD HR',
                        Email='hod.hr@gmail.com',
                        PasswordHash=hash_password('hod123'),
                        DepartmentID=hr_dept.DeptID,
                        RoleID=hod_role.RoleID,
                        IsActive=True,
                        FailedLoginAttempts=0,
                        PasswordLastChanged=datetime.now(IST),
                        PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
                    ),
                    User(  # type: ignore
                        EmpID='EMP004',
                        Username='HOD IT',
                        Email='hod.it@gmail.com',
                        PasswordHash=hash_password('hod123'),
                        DepartmentID=it_dept.DeptID,
                        RoleID=hod_role.RoleID,
                        IsActive=True,
                        FailedLoginAttempts=0,
                        PasswordLastChanged=datetime.now(IST),
                        PasswordExpiryDate=datetime.now(IST) + timedelta(days=90)
                    )
                ]
                db.session.add_all(users)
                db.session.commit()
                print("✓ Test users created:")
                print("  - Admin: emp_id='EMP001', password='admin123' (IT Dept)")
                print("  - Management: emp_id='EMP005', password='manager123' (Approver)")
                print("  - Supervisor: emp_id='EMP006', password='supervisor123' (MIS Consolidator)")
                print("  - HOD Finance: emp_id='EMP002', password='hod123'")
                print("  - HOD HR: emp_id='EMP003', password='hod123'")
                print("  - HOD IT: emp_id='EMP004', password='hod123'")
        
        print("Database initialized successfully!")

def send_monthly_notifications():
    """Automated job to send MIS upload window notifications on the {UPLOAD_WINDOW_START_DAY}st of each month"""
    try:
        with app.app_context():
            hod_role = Role.query.filter_by(RoleName='HOD').first()
            
            if not hod_role:
                logging.error("HOD role not found in database for monthly notification")
                return
            
            hod_users = User.query.filter_by(RoleID=hod_role.RoleID, IsActive=True).all()
            
            if not hod_users:
                logging.warning("No active HOD users found to notify")
                return
            
            if not email_service.is_configured():
                logging.warning("Email service not configured. Skipping monthly notification.")
                return
            
            # Get app URL from environment
            app_url = os.environ.get('REPLIT_DEV_DOMAIN', 'https://your-app.replit.dev')
            if not app_url.startswith('http'):
                app_url = f'https://{app_url}'
            
            success_count, total_count, messages = email_service.send_upload_window_notification(hod_users, app_url)
            
            logging.info(f"Monthly notification sent: {success_count}/{total_count} successful")
            
            for msg in messages:
                if 'error' in msg.lower() or 'failed' in msg.lower():
                    logging.error(f"Notification error: {msg}")
                else:
                    logging.info(f"Notification: {msg}")
    
    except Exception as e:
        logging.error(f"Error in monthly notification job: {str(e)}")

def send_25th_reminder():
    """Send reminder on configured reminder day - final day to upload"""
    try:
        with app.app_context():
            hod_role = Role.query.filter_by(RoleName='HOD').first()
            if not hod_role or not email_service.is_configured():
                return
            
            hod_users = User.query.filter_by(RoleID=hod_role.RoleID, IsActive=True).all()
            
            for user in hod_users:
                subject = f"⚠️ Final Day to Upload MIS - Upload Window Closes on {UPLOAD_WINDOW_REMINDER_DAY}th"
                html_content = f"""
                <!DOCTYPE html>
                <html>
                <body style="font-family: Arial, sans-serif;">
                    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                        <div style="background-color: #f59e0b; color: white; padding: 20px; border-radius: 5px 5px 0 0;">
                            <h2>⚠️ URGENT: Final Day to Upload MIS ({UPLOAD_WINDOW_REMINDER_DAY}th)</h2>
                        </div>
                        <div style="background-color: #f9fafb; padding: 30px; border: 1px solid #e5e7eb;">
                            <p>Dear {user.Username},</p>
                            <div style="background-color: #fef3c7; border-left: 4px solid #f59e0b; padding: 15px; margin: 20px 0;">
                                <strong>This is the FINAL DAY to upload your MIS report!</strong><br>
                                The upload window closes at midnight tonight (11:59 PM).
                            </div>
                            <p><strong>Department:</strong> {user.department.DeptName}</p>
                            <p>Please ensure all monthly reports are submitted before the deadline ends on the {UPLOAD_WINDOW_END_DAY}th.</p>
                            <p>Best regards,<br><strong>MIS System Team</strong></p>
                        </div>
                    </div>
                </body>
                </html>
                """
                email_service.send_email(user.Email, subject, html_content)
            
            logging.info(f"Reminder sent to {len(hod_users)} HOD users on {UPLOAD_WINDOW_REMINDER_DAY}th")
    except Exception as e:
        logging.error(f"Error in reminder: {str(e)}")

def upload_window_lock():
    """Lock uploads on configured lock day - system auto-disable"""
    try:
        with app.app_context():
            logging.info(f"Upload window locked on {UPLOAD_LOCK_DAY}th. Uploads disabled until {UPLOAD_WINDOW_START_DAY}st of next month.")
            # Additional logging or system state updates can be added here
    except Exception as e:
        logging.error(f"Error in upload lock: {str(e)}")

# Initialize scheduler
scheduler = None

def setup_scheduler():
    """Set up automated scheduler for monthly notifications"""
    global scheduler
    
    if scheduler is not None:
        logging.info("Scheduler already running")
        return scheduler
    
    scheduler = BackgroundScheduler(timezone=IST)
    
    # MIS upload window opens
    scheduler.add_job(
        send_monthly_notifications,
        trigger=CronTrigger(day=UPLOAD_WINDOW_START_DAY, hour=UPLOAD_WINDOW_OPEN_HOUR, minute=UPLOAD_WINDOW_OPEN_MINUTE, timezone=IST),
        id='monthly_mis_notification',
        name=f'Send MIS upload window open notification ({UPLOAD_WINDOW_START_DAY}st at {UPLOAD_WINDOW_OPEN_HOUR:02d}:{UPLOAD_WINDOW_OPEN_MINUTE:02d})',
        replace_existing=True
    )
    
    # Final reminder on configured reminder day
    scheduler.add_job(
        send_25th_reminder,
        trigger=CronTrigger(day=UPLOAD_WINDOW_REMINDER_DAY, hour=UPLOAD_WINDOW_REMINDER_HOUR, minute=UPLOAD_WINDOW_REMINDER_MINUTE, timezone=IST),
        id='upload_reminder',
        name=f'Send final day reminder ({UPLOAD_WINDOW_REMINDER_DAY}th at {UPLOAD_WINDOW_REMINDER_HOUR:02d}:{UPLOAD_WINDOW_REMINDER_MINUTE:02d})',
        replace_existing=True
    )
    
    # Upload window lock on configured lock day
    scheduler.add_job(
        upload_window_lock,
        trigger=CronTrigger(day=UPLOAD_LOCK_DAY, hour=UPLOAD_WINDOW_LOCK_HOUR, minute=UPLOAD_WINDOW_LOCK_MINUTE, timezone=IST),
        id='upload_lock',
        name=f'Lock upload window ({UPLOAD_LOCK_DAY}th at {UPLOAD_WINDOW_LOCK_HOUR:02d}:{UPLOAD_WINDOW_LOCK_MINUTE:02d})',
        replace_existing=True
    )
    
    # Supervisor approval window opens on 18th
    def send_supervisor_reminder():
        try:
            with app.app_context():
                supervisor_role = Role.query.filter_by(RoleName='Supervisor').first()
                if supervisor_role and email_service.is_configured():
                    supervisors = User.query.filter_by(RoleID=supervisor_role.RoleID, IsActive=True).all()
                    for supervisor in supervisors:
                        subject = f"Supervisor Approval Window Open - MIS Review Period Begins"
                        html_content = f"""<!DOCTYPE html><html><body style="font-family: Arial;"><div style="max-width: 600px; margin: 0 auto; padding: 20px;"><div style="background-color: #8b5cf6; color: white; padding: 20px;"><h2>MIS Supervisor Approval Window Open</h2></div><div style="background-color: #f9fafb; padding: 30px;"><p>Dear {supervisor.Username},</p><p>The supervisor approval window is now open. Please review all pending HOD MIS uploads and approve them for Management review.</p><p>Access the system to begin reviewing uploads.</p><p>Best regards,<br><strong>MIS System Team</strong></p></div></div></body></html>"""
                        email_service.send_email(supervisor.Email, subject, html_content)
                    logging.info(f"Supervisor approval reminder sent to {len(supervisors)} supervisors")
        except Exception as e:
            logging.error(f"Error in supervisor reminder: {str(e)}")
    
    scheduler.add_job(
        send_supervisor_reminder,
        trigger=CronTrigger(day=SUPERVISOR_APPROVAL_START_DAY, hour=SUPERVISOR_APPROVAL_HOUR, minute=SUPERVISOR_APPROVAL_MINUTE, timezone=IST),
        id='supervisor_approval_reminder',
        name=f'Send supervisor approval window reminder ({SUPERVISOR_APPROVAL_START_DAY}th at {SUPERVISOR_APPROVAL_HOUR:02d}:{SUPERVISOR_APPROVAL_MINUTE:02d})',
        replace_existing=True
    )
    
    scheduler.start()
    logging.info("✓ Scheduler started with 4 automated events:")
    logging.info(f"  - {UPLOAD_WINDOW_START_DAY}st at {UPLOAD_WINDOW_OPEN_HOUR:02d}:{UPLOAD_WINDOW_OPEN_MINUTE:02d}: Upload window opens")
    logging.info(f"  - {SUPERVISOR_APPROVAL_START_DAY}th at {SUPERVISOR_APPROVAL_HOUR:02d}:{SUPERVISOR_APPROVAL_MINUTE:02d}: Supervisor approval window opens")
    logging.info(f"  - {UPLOAD_WINDOW_REMINDER_DAY}th at {UPLOAD_WINDOW_REMINDER_HOUR:02d}:{UPLOAD_WINDOW_REMINDER_MINUTE:02d}: Final reminder")
    logging.info(f"  - {UPLOAD_LOCK_DAY}th at {UPLOAD_WINDOW_LOCK_HOUR:02d}:{UPLOAD_WINDOW_LOCK_MINUTE:02d}: Upload lock")
    
    return scheduler

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    init_db()
    
    # Start automated notification scheduler
    if email_service.is_configured():
        setup_scheduler()
        logging.info("Automated email notifications enabled.")
    else:
        logging.warning("Email service not configured. Automated notifications disabled. Configure SMTP settings to enable.")
    
    try:
        app.run(host='0.0.0.0', port=5000, debug=True)
    finally:
        if scheduler:
            scheduler.shutdown()
            logging.info("Scheduler shut down.")
